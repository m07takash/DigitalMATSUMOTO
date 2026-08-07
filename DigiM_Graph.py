# DigiM_Graph.py
# ============================================================================
# GraphRAG core for Digital MATSUMOTO.
#
# The graph is *pure structure*: Entity nodes + predicate edges. It carries
# no chunk bodies and no free-text descriptions — body text stays in the
# Vector RAG (ChromaDB) side. Both nodes and edges may carry "props"
# (states) such as birthday / residence (node) or role / period (edge).
#
# Storage: one folder per graph (rags master data_type="graph"):
#     <graph_dir>/graph.json       node-link JSON (this module's schema)
#     <graph_dir>/dictionary.json  alias map + seeds + prop_schema
#     <graph_dir>/mapping.json     ingestion source definitions (2 lanes)
#
# Node : { id, name, type, aliases[], domains[],
#          props{key: {value, as_of, source_id, lane}}, embedding? }
# Edge : { source, target, relation, domains[],
#          props{key: {value, as_of, source_id, lane}},
#          source_ids[], create_date }
#
# Retrieval = paths between seed entities (protected) + neighborhoods
# (hop-ascending), capped by EDGE_LIMIT / FANOUT_LIMIT, with DOMAIN_BONUS
# when the query names a domain. Policy knobs live on the agent's
# KNOWLEDGE/BOOK entry, not here and not in the rags master.
# ============================================================================

import os
import csv
import json
import hashlib
import re
from pathlib import Path
from collections import deque

import DigiM_Util as dmu

# Lane priority for prop conflicts: higher wins; ties -> newer as_of wins.
LANE_PRIORITY = {"STRUCTURED": 3, "SEED": 2, "TEXT": 1}

_setting = dmu.read_yaml_file("setting.yaml") or {}
rag_folder_graph_path = _setting.get("RAG_FOLDER_GRAPH", "user/common/rag/graph/")
mst_folder_path = _setting.get("MST_FOLDER", "user/common/mst/")


# ---------------------------------------------------------------- storage --
def resolve_graph_dir(data_name):
    """DATA_NAME -> graph folder. Prefer the rags master (data_type='graph'),
    fall back to RAG_FOLDER_GRAPH/<data_name>/."""
    try:
        rags_file = os.getenv("RAG_MST_FILE") or _setting.get("RAG_MST_FILE", "sample_rags.json")
        rags = dmu.read_json_file(rags_file, mst_folder_path) or {}
        entry = rags.get(data_name) or {}
        if entry.get("data_type") == "graph" and entry.get("file_path"):
            return entry["file_path"]
    except Exception:
        pass
    return str(Path(rag_folder_graph_path) / data_name) + os.sep


def get_graph_list():
    """List available graphs for UI selectors: active rags-master entries
    (data_type='graph') plus RAG_FOLDER_GRAPH subfolders holding a graph.json
    that no master entry already covers."""
    names, covered = [], set()
    try:
        rags_file = os.getenv("RAG_MST_FILE") or _setting.get("RAG_MST_FILE", "sample_rags.json")
        rags = dmu.read_json_file(rags_file, mst_folder_path) or {}
        for k, v in rags.items():
            if v.get("data_type") == "graph" and v.get("active", "Y") == "Y":
                names.append(k)
                covered.add(os.path.normpath(v.get("file_path", "")))
    except Exception:
        pass
    if os.path.isdir(rag_folder_graph_path):
        for d in sorted(os.listdir(rag_folder_graph_path)):
            p = os.path.join(rag_folder_graph_path, d)
            if os.path.exists(os.path.join(p, "graph.json")) and os.path.normpath(p) not in covered:
                names.append(d)
    return names


def load_graph(graph_dir):
    path = str(Path(graph_dir) / "graph.json")
    if os.path.exists(path):
        g = dmu.read_json_file(path) or {}
    else:
        g = {}
    g.setdefault("nodes", {})
    g.setdefault("edges", [])
    return g


def save_graph(graph_dir, graph):
    os.makedirs(graph_dir, exist_ok=True)
    path = str(Path(graph_dir) / "graph.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(graph, f, ensure_ascii=False, indent=1)
    return path


def save_graph_atomic(graph_dir, graph):
    """Same as save_graph but writes to a temp file first and atomically
    renames — protects against a crash mid-write leaving a truncated
    graph.json on disk. Used by the Knowledge Explorer edit UI."""
    os.makedirs(graph_dir, exist_ok=True)
    path = str(Path(graph_dir) / "graph.json")
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(graph, f, ensure_ascii=False, indent=1)
    os.replace(tmp, path)
    return path


def delete_graph(graph_dir):
    """Delete a graph's persisted state — removes `graph.json` only.
    `mapping.json` / `dictionary.json` / `source/` CSVs are preserved so
    the graph can be re-ingested from scratch via DigiM_GraphBuilder or
    generate_rag(). Returns True if the file was removed."""
    path = str(Path(graph_dir) / "graph.json")
    if os.path.exists(path):
        os.remove(path)
        return True
    return False


def filter_active(graph):
    """Return a shallow view of `graph` with nodes/edges marked
    `active == "N"` removed. Missing "active" is treated as active ("Y"),
    so pre-flag graphs work unchanged. Dangling edges (whose endpoints
    got deactivated) are also dropped."""
    _nodes = {nid: n for nid, n in (graph.get("nodes") or {}).items()
              if str(n.get("active", "Y")).upper() != "N"}
    _edges = [e for e in (graph.get("edges") or [])
              if str(e.get("active", "Y")).upper() != "N"
              and e.get("source") in _nodes
              and e.get("target") in _nodes]
    return {"nodes": _nodes, "edges": _edges}


# --------------------------------------------------- CRUD (edit UI helpers) --
def set_node_active(graph, node_id, active=True):
    """Set/clear the logical-delete flag on a node."""
    n = (graph.get("nodes") or {}).get(node_id)
    if n is None:
        return False
    n["active"] = "Y" if active else "N"
    return True


def set_edge_active(graph, edge_index, active=True):
    """Set/clear the logical-delete flag on an edge (by position)."""
    edges = graph.get("edges") or []
    if not (0 <= edge_index < len(edges)):
        return False
    edges[edge_index]["active"] = "Y" if active else "N"
    return True


def update_node(graph, node_id, *, name=None, node_type=None,
                aliases=None, domains=None, active=None):
    """Patch node fields in place. Name change is handled as a rename that
    updates the node dict's `name` field only — the node ID (md5 of the
    ORIGINAL canonical name) stays put so incident edges keep pointing at
    it. Returns True on success, False if the id doesn't exist."""
    n = (graph.get("nodes") or {}).get(node_id)
    if n is None:
        return False
    if name is not None:
        n["name"] = (name or "").strip()
    if node_type is not None:
        n["type"] = node_type
    if aliases is not None:
        n["aliases"] = list(aliases)
    if domains is not None:
        n["domains"] = list(domains)
    if active is not None:
        n["active"] = "Y" if active else "N"
    return True


def update_edge(graph, edge_index, *, relation=None, domains=None,
                source=None, target=None, active=None):
    """Patch edge fields in place. source / target take node ids."""
    edges = graph.get("edges") or []
    if not (0 <= edge_index < len(edges)):
        return False
    e = edges[edge_index]
    if relation is not None:
        e["relation"] = relation
    if domains is not None:
        e["domains"] = list(domains)
    if source is not None:
        e["source"] = source
    if target is not None:
        e["target"] = target
    if active is not None:
        e["active"] = "Y" if active else "N"
    return True


def add_node(graph, name, node_type="", aliases=None, domains=None,
             active=True):
    """Add a brand-new node keyed by md5(name) — like upsert_node but does
    not merge with an existing entry (returns the existing node dict if
    one already exists at that id)."""
    name = (name or "").strip()
    if not name:
        return None
    nid = node_id_for(name)
    if nid in graph.setdefault("nodes", {}):
        return graph["nodes"][nid]
    graph["nodes"][nid] = {
        "id": nid,
        "name": name,
        "type": node_type or "",
        "aliases": list(aliases or []),
        "domains": list(domains or []),
        "props": {},
        "active": "Y" if active else "N",
    }
    return graph["nodes"][nid]


def add_edge(graph, src_id, dst_id, relation, domains=None, props=None,
             create_date="", active=True):
    """Add a brand-new edge (no dedup vs existing edges — the caller
    decides whether duplicate triples are allowed)."""
    if not (src_id in (graph.get("nodes") or {}) and dst_id in (graph.get("nodes") or {})):
        return None
    edge = {
        "source": src_id, "target": dst_id, "relation": relation or "",
        "domains": list(domains or []),
        "props": props or {},
        "source_ids": [],
        "create_date": create_date or "",
        "active": "Y" if active else "N",
    }
    graph.setdefault("edges", []).append(edge)
    return edge


# --------------------------------------------- Excel export / import CRUD --
# Full-round-trip maintenance for the Knowledge Explorer Graph pane.
# Export writes every node/edge (including logically-deleted ones) into a
# two-sheet workbook. Import interprets the workbook as the desired end-state:
#   - existing rows with matching id → UPDATE
#   - rows with blank id            → CREATE
#   - rows with active="N"          → logical delete
#   - rows in the graph but NOT in the workbook → logical delete
# Two-phase apply (plan → confirm → apply) so callers can preview the diff.
def _split_slash_multi(value):
    """Split a `/`-delimited string cell into a list, trimming whitespace.
    Named to avoid collision with the ingestion-side `_split_multi(value, sep)`
    defined further down in this module."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return [p.strip() for p in str(value).split("/") if p.strip()]


def _join_multi(value):
    if not value:
        return ""
    return " / ".join(str(v).strip() for v in value if str(v).strip())


def export_graph_to_xlsx(graph, out_path):
    """Write the graph to `out_path` as a two-sheet .xlsx (nodes / edges).
    Includes active=N rows so a full download reflects the true storage state.
    Edge rows get their list index as the stable `id` column so import can
    match rows back to the original edge."""
    import openpyxl
    wb = openpyxl.Workbook()
    # -- nodes sheet --
    ws_n = wb.active
    ws_n.title = "nodes"
    ws_n.append(["id", "name", "type", "aliases", "domains", "active"])
    for nid, n in (graph.get("nodes") or {}).items():
        ws_n.append([
            nid,
            n.get("name", ""),
            n.get("type", ""),
            _join_multi(n.get("aliases")),
            _join_multi(n.get("domains")),
            "N" if str(n.get("active", "Y")).upper() == "N" else "Y",
        ])
    # -- edges sheet --
    ws_e = wb.create_sheet("edges")
    ws_e.append(["id", "source", "target", "relation", "domains", "active", "create_date"])
    _id_to_name = {nid: (n.get("name") or "") for nid, n in (graph.get("nodes") or {}).items()}
    for ei, e in enumerate(graph.get("edges") or []):
        ws_e.append([
            ei,
            _id_to_name.get(e.get("source"), ""),
            _id_to_name.get(e.get("target"), ""),
            e.get("relation", ""),
            _join_multi(e.get("domains")),
            "N" if str(e.get("active", "Y")).upper() == "N" else "Y",
            e.get("create_date", ""),
        ])
    wb.save(out_path)
    return out_path


def _read_xlsx_rows(xlsx_path):
    """Return {'nodes': [dict, ...], 'edges': [dict, ...]} from the workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = {"nodes": [], "edges": []}
    for sheet in ("nodes", "edges"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                continue
            d = {}
            for k, v in zip(header, row):
                if not k:
                    continue
                d[k] = ("" if v is None else v)
            out[sheet].append(d)
    return out


def compute_graph_import_plan(graph, xlsx_path):
    """Diff the workbook against the graph. Returns a plan dict:
        {"nodes": {"create": [rowdict...], "update": [(nid, changes)...],
                    "delete": [nid, ...]},
         "edges": {"create": [rowdict...], "update": [(ei, changes)...],
                    "delete": [ei, ...]},
         "warnings": [str, ...]}
    Does not mutate the graph — pass the plan to apply_graph_import_plan()."""
    rows = _read_xlsx_rows(xlsx_path)
    warnings = []

    # ---------- nodes ----------
    n_create, n_update, n_delete = [], [], []
    excel_node_ids = set()
    existing_nodes = graph.get("nodes") or {}
    for r in rows["nodes"]:
        # Careful: `int(0) or ""` → "", which would misread edge id 0 as blank.
        _raw_rid = r.get("id")
        rid = "" if _raw_rid is None else str(_raw_rid).strip()
        name = str(r.get("name", "") or "").strip()
        typ = str(r.get("type", "") or "").strip()
        aliases = _split_slash_multi(r.get("aliases"))
        domains = _split_slash_multi(r.get("domains"))
        active = str(r.get("active", "Y") or "Y").strip().upper() != "N"
        if not rid:
            if not name:
                warnings.append(f"[nodes] name 空の行を無視 (id/name のいずれかが必須)")
                continue
            nid_new = node_id_for(name)
            if nid_new in existing_nodes:
                # Treat as update rather than dup create
                n_update.append((nid_new, {
                    "name": name, "type": typ, "aliases": aliases,
                    "domains": domains, "active": active,
                }))
                excel_node_ids.add(nid_new)
            else:
                n_create.append({"name": name, "type": typ, "aliases": aliases,
                                  "domains": domains, "active": active})
            continue
        excel_node_ids.add(rid)
        cur = existing_nodes.get(rid)
        if cur is None:
            warnings.append(f"[nodes] id={rid} は存在しません (削除された ID？) — 無視")
            continue
        changes = {}
        cur_active = str(cur.get("active", "Y")).upper() != "N"
        if (cur.get("name") or "") != name:                 changes["name"] = name
        if (cur.get("type") or "") != typ:                  changes["type"] = typ
        if list(cur.get("aliases") or []) != aliases:       changes["aliases"] = aliases
        if list(cur.get("domains") or []) != domains:       changes["domains"] = domains
        if cur_active != active:                            changes["active"] = active
        if changes:
            n_update.append((rid, changes))
    for nid, n in existing_nodes.items():
        if nid in excel_node_ids:
            continue
        if str(n.get("active", "Y")).upper() == "N":
            # already logically deleted — skip
            continue
        n_delete.append(nid)

    # ---------- edges ----------
    e_create, e_update, e_delete = [], [], []
    excel_edge_ids = set()
    edges = graph.get("edges") or []
    name_to_id = {(nd.get("name") or ""): nid for nid, nd in existing_nodes.items()}
    for r in rows["edges"]:
        # Careful: `int(0) or ""` → "", which would misread edge id 0 as blank.
        _raw_rid = r.get("id")
        rid = "" if _raw_rid is None else str(_raw_rid).strip()
        src_name = str(r.get("source", "") or "").strip()
        tgt_name = str(r.get("target", "") or "").strip()
        relation = str(r.get("relation", "") or "").strip()
        domains = _split_slash_multi(r.get("domains"))
        active = str(r.get("active", "Y") or "Y").strip().upper() != "N"
        create_date = str(r.get("create_date", "") or "").strip()
        src_id = name_to_id.get(src_name)
        tgt_id = name_to_id.get(tgt_name)
        if not rid:
            if not (src_id and tgt_id and relation):
                warnings.append(f"[edges] 新規作成に必要な source/target/relation が不足: {src_name!r} --[{relation}]--> {tgt_name!r}")
                continue
            e_create.append({"source": src_id, "target": tgt_id,
                              "relation": relation, "domains": domains,
                              "active": active, "create_date": create_date})
            continue
        try:
            ei = int(rid)
        except (TypeError, ValueError):
            warnings.append(f"[edges] id={rid!r} が整数ではありません — 無視")
            continue
        if not (0 <= ei < len(edges)):
            warnings.append(f"[edges] id={ei} は範囲外 (現エッジ数 {len(edges)}) — 無視")
            continue
        excel_edge_ids.add(ei)
        cur = edges[ei]
        # Endpoint changes need node-id resolution — skip if unresolvable but
        # note it as a warning (relation-only edits keep working).
        cur_src_name = (existing_nodes.get(cur.get("source")) or {}).get("name", "")
        cur_tgt_name = (existing_nodes.get(cur.get("target")) or {}).get("name", "")
        changes = {}
        if (cur.get("relation") or "") != relation:            changes["relation"] = relation
        if list(cur.get("domains") or []) != domains:          changes["domains"] = domains
        cur_active = str(cur.get("active", "Y")).upper() != "N"
        if cur_active != active:                               changes["active"] = active
        if cur_src_name != src_name:
            if src_id:
                changes["source"] = src_id
            else:
                warnings.append(f"[edges] id={ei}: source={src_name!r} が見つからない — 変更を無視")
        if cur_tgt_name != tgt_name:
            if tgt_id:
                changes["target"] = tgt_id
            else:
                warnings.append(f"[edges] id={ei}: target={tgt_name!r} が見つからない — 変更を無視")
        if changes:
            e_update.append((ei, changes))
    for ei, e in enumerate(edges):
        if ei in excel_edge_ids:
            continue
        if str(e.get("active", "Y")).upper() == "N":
            continue
        e_delete.append(ei)

    return {
        "nodes": {"create": n_create, "update": n_update, "delete": n_delete},
        "edges": {"create": e_create, "update": e_update, "delete": e_delete},
        "warnings": warnings,
    }


def apply_graph_import_plan(graph, plan):
    """Mutate `graph` per the plan produced by compute_graph_import_plan().
    Returns a counts dict: {'nodes_created', 'nodes_updated', 'nodes_deleted',
    'edges_created', 'edges_updated', 'edges_deleted'}."""
    counts = {"nodes_created": 0, "nodes_updated": 0, "nodes_deleted": 0,
              "edges_created": 0, "edges_updated": 0, "edges_deleted": 0}

    for row in plan["nodes"]["create"]:
        n = add_node(graph, row["name"], node_type=row.get("type", ""),
                      aliases=row.get("aliases"), domains=row.get("domains"),
                      active=row.get("active", True))
        if n is not None:
            counts["nodes_created"] += 1
    for nid, changes in plan["nodes"]["update"]:
        _kw = {}
        if "name" in changes:    _kw["name"] = changes["name"]
        if "type" in changes:    _kw["node_type"] = changes["type"]
        if "aliases" in changes: _kw["aliases"] = changes["aliases"]
        if "domains" in changes: _kw["domains"] = changes["domains"]
        if "active" in changes:  _kw["active"] = changes["active"]
        if update_node(graph, nid, **_kw):
            counts["nodes_updated"] += 1
    for nid in plan["nodes"]["delete"]:
        if set_node_active(graph, nid, False):
            counts["nodes_deleted"] += 1

    for row in plan["edges"]["create"]:
        e = add_edge(graph, row["source"], row["target"], row["relation"],
                      domains=row.get("domains"),
                      create_date=row.get("create_date", ""),
                      active=row.get("active", True))
        if e is not None:
            counts["edges_created"] += 1
    for ei, changes in plan["edges"]["update"]:
        if update_edge(graph, ei, **{k: v for k, v in changes.items() if k in
                                      ("relation", "domains", "source", "target", "active")}):
            counts["edges_updated"] += 1
    for ei in plan["edges"]["delete"]:
        if set_edge_active(graph, ei, False):
            counts["edges_deleted"] += 1

    return counts


def load_dictionary(graph_dir):
    path = str(Path(graph_dir) / "dictionary.json")
    d = dmu.read_json_file(path) if os.path.exists(path) else {}
    d = d or {}
    d.setdefault("aliases", {})
    d.setdefault("seeds", [])
    d.setdefault("prop_schema", {})
    return d


# ------------------------------------------------------- entity resolution --
def canonical_name(name, dictionary):
    """Normalize an entity mention through the alias map."""
    name = (name or "").strip()
    return dictionary.get("aliases", {}).get(name, name)


def node_id_for(name):
    """Deterministic node id from the canonical name (stable across builds)."""
    return "ent_" + hashlib.md5(name.encode("utf-8")).hexdigest()[:12]


def upsert_node(graph, name, node_type="", aliases=None, domains=None, dictionary=None):
    """Create or merge a node by canonical name. Returns the node dict."""
    dictionary = dictionary or {"aliases": {}}
    cname = canonical_name(name, dictionary)
    if not cname:
        return None
    nid = node_id_for(cname)
    node = graph["nodes"].get(nid)
    if node is None:
        node = {"id": nid, "name": cname, "type": node_type or "",
                "aliases": [], "domains": [], "props": {}}
        graph["nodes"][nid] = node
    if node_type and not node.get("type"):
        node["type"] = node_type
    for a in (aliases or []):
        a = a.strip()
        if a and a != cname and a not in node["aliases"]:
            node["aliases"].append(a)
    for d in (domains or []):
        if d and d not in node["domains"]:
            node["domains"].append(d)
    return node


def set_prop(props, key, value, as_of="", source_id="", lane="STRUCTURED"):
    """Overwrite-semantics state write. Higher lane wins; same lane -> newer
    as_of wins (missing as_of loses to any dated value)."""
    if value in (None, ""):
        return
    cur = props.get(key)
    if cur:
        cur_p = LANE_PRIORITY.get(cur.get("lane", "TEXT"), 1)
        new_p = LANE_PRIORITY.get(lane, 1)
        if new_p < cur_p:
            return
        if new_p == cur_p and str(as_of or "") < str(cur.get("as_of") or ""):
            return
    props[key] = {"value": value, "as_of": as_of or "", "source_id": source_id or "", "lane": lane}


def upsert_edge(graph, src_id, dst_id, relation, domains=None, props=None,
                source_id="", create_date="", lane="STRUCTURED"):
    """Create or merge an edge keyed by (source, target, relation)."""
    if not (src_id and dst_id and relation) or src_id == dst_id:
        return None
    edge = None
    for e in graph["edges"]:
        if e["source"] == src_id and e["target"] == dst_id and e["relation"] == relation:
            edge = e
            break
    if edge is None:
        edge = {"source": src_id, "target": dst_id, "relation": relation,
                "domains": [], "props": {}, "source_ids": [], "create_date": create_date or ""}
        graph["edges"].append(edge)
    for d in (domains or []):
        if d and d not in edge["domains"]:
            edge["domains"].append(d)
    for k, v in (props or {}).items():
        if isinstance(v, dict):
            set_prop(edge["props"], k, v.get("value"), v.get("as_of", ""), v.get("source_id", ""), v.get("lane", lane))
        else:
            set_prop(edge["props"], k, v, create_date, source_id, lane)
    if source_id and source_id not in edge["source_ids"]:
        edge["source_ids"].append(source_id)
    if create_date and (not edge["create_date"] or create_date > edge["create_date"]):
        edge["create_date"] = create_date
    return edge


# ------------------------------------------------------------- ingestion ---
def _split_multi(value, sep):
    return [v.strip() for v in str(value or "").split(sep) if v.strip()]


def ingest_seeds(graph, dictionary):
    """Dictionary seeds become nodes unconditionally (lane=SEED)."""
    for seed in dictionary.get("seeds", []):
        upsert_node(graph, seed.get("name", ""), seed.get("type", ""),
                    aliases=seed.get("aliases", []), domains=seed.get("domains", []),
                    dictionary=dictionary)


def ingest_structured_source(graph, dictionary, graph_dir, source, sep=";"):
    """Lane A: deterministic column mapping. No LLM involved."""
    mapping = source.get("MAPPING", {})
    file_path = str(Path(graph_dir) / source.get("FILE", ""))
    if not os.path.exists(file_path):
        return {"file": source.get("FILE"), "rows": 0, "error": "file not found"}

    ent_map = mapping.get("ENTITY", {})
    props_map = mapping.get("PROPS", {})
    rel_maps = mapping.get("RELATIONS", [])
    domains_col = mapping.get("DOMAINS", "")
    as_of_col = mapping.get("AS_OF", "")

    rows = 0
    with open(file_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            rows += 1
            name = row.get(ent_map.get("name", ""), "")
            if not name:
                continue
            # ENTITY.type is either a literal type label or a column name.
            type_spec = ent_map.get("type", "")
            node_type = row.get(type_spec, type_spec)
            aliases = _split_multi(row.get(ent_map.get("aliases", ""), ""), sep)
            domains = _split_multi(row.get(domains_col, ""), sep)
            as_of = row.get(as_of_col, "")

            node = upsert_node(graph, name, node_type, aliases, domains, dictionary)
            if node is None:
                continue
            for prop_key, col in props_map.items():
                set_prop(node["props"], prop_key, row.get(col, ""), as_of,
                         source.get("FILE", ""), "STRUCTURED")

            for rel in rel_maps:
                targets = _split_multi(row.get(rel.get("target_col", ""), ""), sep)
                for target_name in targets:
                    tnode = upsert_node(graph, target_name, "", [], domains, dictionary)
                    if tnode is None:
                        continue
                    edge_props = {k: row.get(col, "") for k, col in (rel.get("props") or {}).items()}
                    # direction OUT (default): entity -> target / IN: target -> entity
                    if rel.get("direction", "OUT") == "IN":
                        src, dst = tnode["id"], node["id"]
                    else:
                        src, dst = node["id"], tnode["id"]
                    upsert_edge(graph, src, dst, rel.get("relation", "関連"),
                                domains=domains, props=edge_props,
                                source_id=source.get("FILE", ""), create_date=as_of,
                                lane="STRUCTURED")
    return {"file": source.get("FILE"), "rows": rows}


GRAPH_EXTRACT_PROMPT = """あなたは知識グラフの抽出器です。以下の本文から三つ組と状態を抽出し、JSONのみで出力してください。

ルール:
- 述語(relation)は「関連」のような曖昧語を避け、評価/懸念/提案/参画/策定/対照 などの具体動詞にする
- エンティティは (1)他エンティティとも関係を持ちうる (2)質問の主題になりうる もののみ。数値・日付・役職などは props に入れる
- 出力形式: {"triples": [{"subject": "", "subject_type": "", "relation": "", "object": "", "object_type": "", "props": {"期間": "", "役割": ""}}], "node_props": [{"entity": "", "key": "", "value": ""}]}

既知のエンティティ: %s

本文:
%s
"""


def ingest_text_source(graph, dictionary, graph_dir, source, llm_extractor, sep=";"):
    """Lane B: free-text extraction through an LLM extractor callable.
    `llm_extractor(prompt) -> str(JSON)` keeps this module offline-testable."""
    mapping = source.get("MAPPING", {})
    file_path = str(Path(graph_dir) / source.get("FILE", ""))
    if not os.path.exists(file_path):
        return {"file": source.get("FILE"), "rows": 0, "error": "file not found"}
    if llm_extractor is None:
        return {"file": source.get("FILE"), "rows": 0, "skipped": "no LLM extractor (use --use-llm)"}

    text_col = mapping.get("TEXT", "text")
    domains_col = mapping.get("DOMAINS", "category")
    as_of_col = mapping.get("AS_OF", "create_date")
    source_id_col = mapping.get("SOURCE_ID", "")

    known = [n["name"] for n in graph["nodes"].values()]
    rows = 0
    with open(file_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            rows += 1
            text = row.get(text_col, "")
            if not text:
                continue
            domains = _split_multi(row.get(domains_col, ""), sep)
            as_of = row.get(as_of_col, "")
            source_id = row.get(source_id_col, "") or source.get("FILE", "")

            prompt = GRAPH_EXTRACT_PROMPT % (", ".join(known[:50]), text)
            try:
                raw = llm_extractor(prompt)
                data = json.loads(re.sub(r"^```(json)?|```$", "", str(raw).strip(), flags=re.M))
            except Exception as e:
                data = {}
                print(f"[DigiM_Graph] extract failed ({source_id}): {e}")
            for t in data.get("triples", []):
                s = upsert_node(graph, t.get("subject", ""), t.get("subject_type", ""), [], domains, dictionary)
                o = upsert_node(graph, t.get("object", ""), t.get("object_type", ""), [], domains, dictionary)
                if s and o:
                    upsert_edge(graph, s["id"], o["id"], t.get("relation", "関連"),
                                domains=domains, props=t.get("props") or {},
                                source_id=source_id, create_date=as_of, lane="TEXT")
            for p in data.get("node_props", []):
                node = upsert_node(graph, p.get("entity", ""), "", [], domains, dictionary)
                if node:
                    set_prop(node["props"], p.get("key", ""), p.get("value", ""), as_of, source_id, "TEXT")
    return {"file": source.get("FILE"), "rows": rows}


def promote_props_to_edges(graph, dictionary):
    """Props whose value matches an existing node (or seed) become edges.
    Deterministic: dictionary/name match only. e.g. 居住地:東京 -> --[居住地]--> (東京)
    only when 東京 is already a node."""
    name_index = {n["name"]: n["id"] for n in graph["nodes"].values()}
    for n in list(graph["nodes"].values()):
        for key in list(n["props"].keys()):
            p = n["props"][key]
            cval = canonical_name(str(p.get("value", "")), dictionary)
            target_id = name_index.get(cval)
            if target_id and target_id != n["id"]:
                upsert_edge(graph, n["id"], target_id, key,
                            domains=n.get("domains", []),
                            source_id=p.get("source_id", ""), create_date=p.get("as_of", ""),
                            lane=p.get("lane", "STRUCTURED"))
                del n["props"][key]


def embed_nodes(graph):
    """Optional: node embeddings from 'name (type) [domains] aliases'.
    Requires an embedding API key; failures leave nodes without embeddings
    (retrieval then falls back to lexical linking)."""
    done, failed = 0, 0
    for n in graph["nodes"].values():
        base = f'{n["name"]} ({n.get("type","")}) [{" ".join(n.get("domains",[]))}] {" ".join(n.get("aliases",[]))}'
        try:
            vec = dmu.embed_text(base)
            if vec:
                n["embedding"] = vec
                done += 1
        except Exception:
            failed += 1
    return {"embedded": done, "failed": failed}


def build_graph(graph_dir, use_llm=False, embed=False, llm_agent_file="agent_67GraphExtract.json"):
    """Full build from mapping.json + dictionary.json. Returns a report dict."""
    mapping = dmu.read_json_file(str(Path(graph_dir) / "mapping.json")) or {}
    dictionary = load_dictionary(graph_dir)
    sep = mapping.get("MULTI_VALUE_SEPARATOR", ";")
    graph = {"nodes": {}, "edges": []}

    report = {"graph_name": mapping.get("GRAPH_NAME", ""), "sources": []}
    ingest_seeds(graph, dictionary)

    llm_extractor = None
    if use_llm:
        import DigiM_Agent as dma
        def llm_extractor(prompt):
            # ext_generate_pureLLM returns (response, model_name, prompt_tokens,
            # response_tokens); ingest_text_source expects the raw JSON string.
            _ret = dma.ext_generate_pureLLM(llm_agent_file, prompt)
            if isinstance(_ret, tuple) and _ret:
                return _ret[0]
            return _ret

    for source in mapping.get("SOURCES", []):
        lane = source.get("LANE", "STRUCTURED")
        if lane == "STRUCTURED":
            report["sources"].append(ingest_structured_source(graph, dictionary, graph_dir, source, sep))
        else:
            report["sources"].append(ingest_text_source(graph, dictionary, graph_dir, source, llm_extractor, sep))

    promote_props_to_edges(graph, dictionary)
    if embed:
        report["embedding"] = embed_nodes(graph)

    report["nodes"] = len(graph["nodes"])
    report["edges"] = len(graph["edges"])
    report["path"] = save_graph(graph_dir, graph)
    return report


# -------------------------------------------------------------- retrieval --
def _adjacency(graph):
    adj = {}
    for i, e in enumerate(graph["edges"]):
        adj.setdefault(e["source"], []).append(i)
        adj.setdefault(e["target"], []).append(i)
    return adj


def link_entities(query, graph, dictionary, query_vecs=None, top_k=4,
                    sim_threshold=0.45, trace=None):
    """Seed selection: lexical first (aliases + names as substrings of the
    query), then embedding similarity when node embeddings and query vectors
    are both available.

    If `trace` is passed (a mutable dict), the function records:
      trace["matches_raw"]        : [{mention, mapped_to_name}] before alias resolution
      trace["matches_normalized"] : [{node_id, node_name}] after alias resolution
      trace["source"]             : "lexical" | "embedding"
    Used by the seed-provenance display in Detail Information / Analytics."""
    seeds = {}
    mentions = dict(dictionary.get("aliases", {}))
    for n in graph["nodes"].values():
        mentions[n["name"]] = n["name"]
        for a in n.get("aliases", []):
            mentions.setdefault(a, n["name"])
    for mention, cname in mentions.items():
        if mention and mention in query:
            cn = canonical_name(cname, dictionary)
            nid = node_id_for(cn)
            if nid in graph["nodes"]:
                _prev = seeds.get(nid, 0.0)
                seeds[nid] = max(_prev, 1.0)
                if trace is not None:
                    trace.setdefault("matches_raw", []).append(
                        {"mention": mention, "mapped_to_name": cn})
                    if _prev == 0.0:
                        trace.setdefault("matches_normalized", []).append(
                            {"node_id": nid,
                             "node_name": graph["nodes"][nid]["name"]})

    if query_vecs:
        scored = []
        for n in graph["nodes"].values():
            vec = n.get("embedding")
            if not vec:
                continue
            best = max((dmu.calculate_similarity_vec(qv, vec, "Cosine") for qv in query_vecs if qv), default=0.0)
            if best >= sim_threshold:
                scored.append((best, n["id"]))
        for best, nid in sorted(scored, reverse=True)[:top_k]:
            seeds[nid] = max(seeds.get(nid, 0.0), round(best, 3))
    return seeds  # {node_id: confidence}


def link_entities_multi(queries, graph, dictionary, query_vecs=None,
                          top_k=4, sim_threshold=0.45):
    """Union `link_entities` results across multiple queries. Returns
    (seeds_dict, seed_trace) where seed_trace lists per-query provenance
    (raw substring match → normalized node) suitable for the Detail /
    Analytics seed-provenance panels. queries may be a single string or
    an iterable of strings."""
    if isinstance(queries, str):
        queries = [queries]
    queries = [str(q) for q in (queries or []) if str(q).strip()]
    seeds = {}
    seed_trace = []
    for q in queries:
        trace = {"query": q, "matches_raw": [], "matches_normalized": []}
        _s = link_entities(q, graph, dictionary, query_vecs=query_vecs,
                             top_k=top_k, sim_threshold=sim_threshold, trace=trace)
        for _nid, _conf in _s.items():
            seeds[_nid] = max(seeds.get(_nid, 0.0), _conf)
        seed_trace.append(trace)
    return seeds, seed_trace


def detect_domains(query, graph):
    """Domains explicitly named in the query -> boost set."""
    all_domains = set()
    for n in graph["nodes"].values():
        all_domains.update(n.get("domains", []))
    for e in graph["edges"]:
        all_domains.update(e.get("domains", []))
    return {d for d in all_domains if d and d in query}


def _shortest_path_edges(graph, adj, src, dst, max_len):
    """BFS shortest path (undirected view). Returns list of edge indexes."""
    if src == dst:
        return []
    prev = {src: (None, None)}
    q = deque([(src, 0)])
    while q:
        cur, depth = q.popleft()
        if depth >= max_len:
            continue
        for ei in adj.get(cur, []):
            e = graph["edges"][ei]
            nxt = e["target"] if e["source"] == cur else e["source"]
            if nxt in prev:
                continue
            prev[nxt] = (cur, ei)
            if nxt == dst:
                path = []
                node = dst
                while prev[node][0] is not None:
                    path.append(prev[node][1])
                    node = prev[node][0]
                return list(reversed(path))
            q.append((nxt, depth + 1))
    return None


def _edge_score(edge, boost_domains, domain_bonus):
    score = 1.0
    if boost_domains and set(edge.get("domains", [])) & boost_domains:
        score *= domain_bonus
    return score


def search_graph(query, graph_dir, policy, query_vecs=None, meta_dates=None,
                   queries=None):
    """Core retrieval: paths between seeds (protected) + hop-ascending
    neighborhoods, capped by EDGE_LIMIT / FANOUT_LIMIT.
    Returns dict with seeds / paths / edges (selected) / domains / seed_trace.
    Logically-deleted entries (active="N") are filtered out up-front.

    When `queries` (list of str) is provided, seed detection runs on ALL
    queries and the results are unioned — used when the caller has an
    LLM-expanded / digest-augmented query alongside the raw user query.
    Falls back to `[query]` when queries is omitted."""
    raw = load_graph(graph_dir)
    graph = filter_active(raw)
    dictionary = load_dictionary(graph_dir)
    adj = _adjacency(graph)

    hops = int(policy.get("HOPS", 2))
    edge_limit = int(policy.get("EDGE_LIMIT", 30))
    fanout_limit = int(policy.get("FANOUT_LIMIT", 5))
    domain_bonus = float(policy.get("DOMAIN_BONUS", 1.3))

    _all_queries = list(queries) if queries else [query]
    seeds, seed_trace = link_entities_multi(_all_queries, graph, dictionary, query_vecs)
    boost_domains = set()
    for _q in _all_queries:
        boost_domains |= detect_domains(_q, graph)

    selected = []          # [(edge_index, hop, kind)]
    selected_set = set()

    # 1) Paths between every seed pair (protected allocation).
    seed_ids = list(seeds.keys())
    paths = []
    for i in range(len(seed_ids)):
        for j in range(i + 1, len(seed_ids)):
            p = _shortest_path_edges(graph, adj, seed_ids[i], seed_ids[j], max_len=hops * 2)
            if p:
                paths.append(p)
                for ei in p:
                    if ei not in selected_set and len(selected) < edge_limit:
                        selected.append((ei, 0, "path"))
                        selected_set.add(ei)

    # 2) Neighborhood expansion, hop by hop, with per-node fanout pruning.
    frontier = set(seed_ids)
    visited = set(seed_ids)
    for hop in range(1, hops + 1):
        next_frontier = set()
        for nid in frontier:
            incident = adj.get(nid, [])
            # Rank: highest edge_score first (domain-boosted); tie -> NEWEST
            # create_date first (recent memos usually reflect the current
            # question better than aged ones). Empty create_date is treated
            # as very old ("") so it sorts last after real dates.
            ranked = sorted(
                incident,
                key=lambda ei: (_edge_score(graph["edges"][ei], boost_domains, domain_bonus),
                                graph["edges"][ei].get("create_date", "")),
                reverse=True,
            )[:fanout_limit]
            for ei in ranked:
                e = graph["edges"][ei]
                other = e["target"] if e["source"] == nid else e["source"]
                if ei not in selected_set and len(selected) < edge_limit:
                    selected.append((ei, hop, "neighbor"))
                    selected_set.add(ei)
                if other not in visited:
                    next_frontier.add(other)
                    visited.add(other)
        frontier = next_frontier
        if len(selected) >= edge_limit:
            break

    return {"graph": graph, "seeds": seeds, "paths": paths,
            "selected": selected, "boost_domains": boost_domains,
            "seed_trace": seed_trace, "queries": _all_queries}


# ----------------------------------------------------- context assembly ----
def _props_inline(props, max_items=3):
    items = [f'{k}:{v.get("value","")}' for k, v in list(props.items())[:max_items] if v.get("value")]
    return f' ({", ".join(items)})' if items else ""


def _render_edge(graph, edge, template):
    src = graph["nodes"].get(edge["source"], {}).get("name", edge["source"])
    dst = graph["nodes"].get(edge["target"], {}).get("name", edge["target"])
    date = f' ({edge.get("create_date","")[:7]})' if edge.get("create_date") else ""
    return template.format(subject=src, relation=edge["relation"] + _props_inline(edge.get("props", {})),
                           object=dst, description="", date=date)


# ------------------------------------------------------- usage analysis ----
def link_output(text, graph, dictionary, min_mention_len=2):
    """Match a *generated response* back onto the graph (lexical, zero-API).

    Returns:
      nodes : {node_id: mention_count}   (name + alias occurrences summed)
      edges : [{"index": i, "freq": n, "predicate_match": bool}]
              An edge counts as "used in the output" when BOTH endpoint
              nodes are mentioned (co-occurrence rule). freq is the weaker
              endpoint's count; predicate_match flags when the relation
              string itself also appears in the text.
    """
    text = text or ""
    node_counts = {}
    mention_map = {}  # mention string -> node_id
    for n in graph["nodes"].values():
        mention_map[n["name"]] = n["id"]
        for a in n.get("aliases", []):
            mention_map.setdefault(a, n["id"])
    for mention, cname in dictionary.get("aliases", {}).items():
        nid = node_id_for(cname)
        if nid in graph["nodes"]:
            mention_map.setdefault(mention, nid)

    for mention, nid in mention_map.items():
        if len(mention) < min_mention_len:
            continue
        c = text.count(mention)
        if c:
            node_counts[nid] = node_counts.get(nid, 0) + c

    edge_hits = []
    for i, e in enumerate(graph["edges"]):
        cs = node_counts.get(e["source"], 0)
        ct = node_counts.get(e["target"], 0)
        if cs and ct:
            edge_hits.append({
                "index": i,
                "freq": min(cs, ct),
                "predicate_match": bool(e.get("relation")) and e["relation"] in text,
            })
    return {"nodes": node_counts, "edges": edge_hits}


def analyze_graph_usage(query, response_text, rag, query_vecs=None, queries=None):
    """On-demand per-turn usage analysis for Knowledge Utility.

    Recomputes the (deterministic, lexical) retrieval for `query` and links
    `response_text` back onto the graph, then diffs the two:

      seeds           : query-linked seed node ids
      retrieved_nodes : node ids touched by the retrieval (seeds + endpoints)
      retrieved_edges : edge indexes selected by the retrieval
      output_nodes    : {node_id: mention count in the response}
      output_edges    : [{"index", "freq", "predicate_match"}]
      missed_nodes    : in the output but NOT retrieved  (coverage gap = red)
      missed_edges    : same, for edges

    Needs only the session log's (query, response) pair — no runtime
    recording. Note: results reflect the *current* graph; if the graph was
    rebuilt since the turn, the replayed retrieval may differ slightly.
    """
    data_list = [d for d in rag.get("DATA", []) if d.get("DATA_TYPE") == "GRAPH"]
    if not data_list:
        return None
    graph_dir = resolve_graph_dir(data_list[0]["DATA_NAME"])
    result = search_graph(query, graph_dir, rag, query_vecs=query_vecs,
                           queries=queries)
    graph = result["graph"]
    dictionary = load_dictionary(graph_dir)

    retrieved_edges = {ei for ei, _hop, _kind in result["selected"]}
    retrieved_nodes = set(result["seeds"].keys())
    for ei in retrieved_edges:
        e = graph["edges"][ei]
        retrieved_nodes.add(e["source"])
        retrieved_nodes.add(e["target"])

    out = link_output(response_text, graph, dictionary)
    missed_nodes = [nid for nid in out["nodes"] if nid not in retrieved_nodes]
    missed_edges = [h for h in out["edges"] if h["index"] not in retrieved_edges]

    return {
        "graph": graph,
        "graph_dir": graph_dir,
        "seeds": list(result["seeds"].keys()),
        "seed_names": [
            graph["nodes"].get(nid, {}).get("name", nid)
            for nid in result["seeds"].keys()
        ],
        "seed_trace": result.get("seed_trace", []),
        "queries": result.get("queries", [query]),
        "paths": result["paths"],
        "retrieved_nodes": sorted(retrieved_nodes),
        "retrieved_edges": sorted(retrieved_edges),
        "output_nodes": out["nodes"],
        "output_edges": out["edges"],
        "missed_nodes": missed_nodes,
        "missed_edges": missed_edges,
        "boost_domains": sorted(result["boost_domains"]),
    }


def build_graph_context(query, rag, exec_info=None, query_vecs=None, meta_searches=None):
    """Entry point used by DigiM_Context.create_rag_context (RETRIEVER='Graph').
    Returns (rag_context, rag_selected) in the same shape the Vector /
    PageIndex paths produce: context string + list of log-format strings.

    When `exec_info["_QUERIES"]` is a list of strings (see the multi-query
    RAG generator path), seed detection runs on ALL queries and unions the
    result. The full seed trace (pre/post alias mapping, per query) is
    stashed back into `exec_info["_GRAPH_SEED_TRACES"]` (keyed by RAG_NAME)
    so downstream Detail Information / Analytics can display seed
    provenance without re-running the retrieval."""
    data_list = [d for d in rag.get("DATA", []) if d.get("DATA_TYPE") == "GRAPH"]
    if not data_list:
        return "", []

    header = rag.get("HEADER_TEMPLATE", "以下はあなたの知識グラフから取得した関連構造です。\n")
    path_tpl = rag.get("PATH_TEMPLATE", "・{path}\n")
    edge_tpl = rag.get("EDGE_TEMPLATE", "・({subject}) --[{relation}]--> ({object}){date}\n")
    log_tpl = rag.get("LOG_TEMPLATE",
        "'rag':'{rag_name}', 'DB': 'Graph', 'QUERY_SEQ': '{query_seq}', 'QUERY_MODE': '{query_mode}', "
        "'ID': '{id}', 'similarity_Q': {similarity_prompt}, 'similarity_A': {similarity_response}, "
        "'relation': '{relation}', 'text_short': '{text_short}'")
    props_limit = int(rag.get("PROPS_LIMIT", 5))

    _exec = exec_info if isinstance(exec_info, dict) else {}
    _multi_queries = _exec.get("_QUERIES") or [query]
    if isinstance(_multi_queries, str):
        _multi_queries = [_multi_queries]

    rag_context = ""
    rag_selected = []

    for rd in data_list:
        graph_dir = resolve_graph_dir(rd["DATA_NAME"])
        result = search_graph(query, graph_dir, rag, query_vecs=query_vecs,
                                queries=_multi_queries)
        # Stash seed trace for downstream Detail / Analytics rendering.
        if isinstance(exec_info, dict):
            _traces = exec_info.setdefault("_GRAPH_SEED_TRACES", {})
            _traces[rag.get("RAG_NAME", "")] = {
                "data_name": rd.get("DATA_NAME", ""),
                "queries": _multi_queries,
                "seed_trace": result.get("seed_trace", []),
                "final_seed_names": [
                    result["graph"]["nodes"].get(nid, {}).get("name", nid)
                    for nid in (result.get("seeds") or {}).keys()
                ],
            }
        graph = result["graph"]
        if not result["selected"]:
            continue

        query_mode = "(GRAPH:local"
        if result["boost_domains"]:
            query_mode += "+domain:" + "/".join(sorted(result["boost_domains"]))
        query_mode += ")"

        block = header

        # === Path === follow BFS edge order, orienting each edge relative to the
        # previous node in the walk: (A) --[r]--> (B) <--[r]-- (C) …
        path_lines = []
        for p in result["paths"]:
            walk = []
            prev_node = None
            for idx, ei in enumerate(p):
                e = graph["edges"][ei]
                sname = graph["nodes"].get(e["source"], {}).get("name", "")
                dname = graph["nodes"].get(e["target"], {}).get("name", "")
                if idx == 0:
                    # find shared node with next edge to determine start
                    if len(p) > 1:
                        e2 = graph["edges"][p[1]]
                        shared = {e["source"], e["target"]} & {e2["source"], e2["target"]}
                        start = ({e["source"], e["target"]} - shared).pop() if shared else e["source"]
                    else:
                        start = e["source"]
                    prev_node = start
                    walk.append(f"({graph['nodes'].get(start, {}).get('name', '')})")
                rel = e["relation"] + _props_inline(e.get("props", {}))
                if e["source"] == prev_node:
                    walk.append(f" --[{rel}]--> ({dname})")
                    prev_node = e["target"]
                else:
                    walk.append(f" <--[{rel}]-- ({sname})")
                    prev_node = e["source"]
            path_lines.append(path_tpl.format(path="".join(walk)))
        if path_lines:
            block += "■経路（質問に関わるつながり）\n" + "".join(dict.fromkeys(path_lines))

        # === Relations (neighbouring edges outside the path) ===
        rel_lines = []
        for ei, hop, kind in result["selected"]:
            if kind == "path":
                continue
            rel_lines.append(_render_edge(graph, graph["edges"][ei], edge_tpl))
        if rel_lines:
            block += "■関係（近傍・関連度順）\n" + "".join(rel_lines)

        # === States (seed nodes only) ===
        state_lines = []
        for nid in result["seeds"]:
            node = graph["nodes"].get(nid, {})
            props = {k: v for k, v in node.get("props", {}).items() if v.get("value")}
            if props:
                items = [f'{k}={v["value"]}' for k, v in list(props.items())[:props_limit]]
                state_lines.append(f'・{node.get("name","")}: {", ".join(items)}\n')
        if state_lines:
            block += "■主要エンティティの状態\n" + "".join(state_lines)

        rag_context += block + "\n"

        # References (legacy string log entries, PageIndex-style).
        for ei, hop, kind in result["selected"]:
            e = graph["edges"][ei]
            sname = graph["nodes"].get(e["source"], {}).get("name", "")
            dname = graph["nodes"].get(e["target"], {}).get("name", "")
            try:
                rag_selected.append(log_tpl.format(
                    rag_name=rag.get("RAG_NAME", ""),
                    query_seq="0", query_mode=query_mode,
                    id=f'{e["source"]}->{e["target"]}',
                    similarity_prompt=0.0, similarity_response=0.0,
                    relation=e["relation"],
                    text_short=f'({sname}) --[{e["relation"]}]--> ({dname})'[:50],
                ))
            except (KeyError, IndexError):
                rag_selected.append(f"'rag':'{rag.get('RAG_NAME','')}', 'QUERY_MODE': '{query_mode}', "
                                    f"'text_short': '({sname}) --[{e['relation']}]--> ({dname})'")

    return rag_context, rag_selected
