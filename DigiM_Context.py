import os
import shutil
import csv
import json
import ast
import logging
import threading
import pandas as pd
import sqlite3
import chromadb
from chromadb.errors import NotFoundError
import re
import mimetypes
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor

import DigiM_Util as dmu

logger = logging.getLogger(__name__)
import DigiM_Tool as dmt
import DigiM_Notion as dmn

# Load folder paths and other settings from setting.yaml
system_setting_dict = dmu.read_yaml_file("setting.yaml")
mst_folder_path = system_setting_dict["MST_FOLDER"]
temp_folder_path = system_setting_dict["TEMP_FOLDER"]
rag_folder_db_path = system_setting_dict["RAG_FOLDER_DB"]
rag_folder_pages_path = system_setting_dict.get("RAG_FOLDER_PAGES", "user/common/rag/pages/")
agent_folder_path = system_setting_dict.get("AGENT_FOLDER", "user/common/agent/")

# Load system.env and set environment variables
if os.path.exists("system.env"):
    load_dotenv("system.env")
prompt_template_mst_file = os.getenv("PROMPT_TEMPLATE_MST_FILE")
prompt_temp_mst_path = str(Path(mst_folder_path) / prompt_template_mst_file)
rag_mst_file = os.getenv("RAG_MST_FILE")
notion_db_mst_file = os.getenv("NOTION_MST_FILE")

# B-1: Singleton ChromaDB client (thread-safe)
_chroma_client = None
_chroma_client_lock = threading.Lock()

def get_chroma_client():
    global _chroma_client
    if _chroma_client is None:
        with _chroma_client_lock:
            if _chroma_client is None:  # double-checked locking
                _chroma_client = chromadb.PersistentClient(path=rag_folder_db_path)
    return _chroma_client

# Build context from attached content
def create_contents_context(agent_data, contents, seq=0, sub_seq=0):
    contents_context = ""
    contents_records = []
    image_files = []
    file_seq = 0

    for content in contents:
        content_context, content_record, image_file = get_text_content(agent_data, content, seq, sub_seq, file_seq)
        contents_context += content_context
        contents_records.append(content_record)
        if image_file:
            if isinstance(image_file, list):
                image_files.extend(image_file)
            else:
                image_files.append(image_file)
        file_seq += 1

    return contents_context, contents_records, image_files

# Extract content text from an uploaded file
def get_text_content(agent_data, content, seq, sub_seq, file_seq):
    content_context = ""
    image_file = ""

    # Set up content-related info
    if os.path.basename(content).startswith("[IN]") or os.path.basename(content).startswith("[OUT]"):
        file_name = os.path.basename(content)
    else:
        file_name = "[IN]seq"+str(seq)+"-"+str(sub_seq)+"_"+str(file_seq)+"_"+os.path.basename(content)
    file_size = os.path.getsize(content)
    file_type, encoding = mimetypes.guess_type(content)

    # Configure the support agent
    support_agent = agent_data["SUPPORT_AGENT"]

    # Read the file extension
    file_ext = os.path.splitext(content)[1].lower()
    _header = f"<br>---------<br>File name: {file_name}<br><br>"

    # Build context per file type
    if file_ext in [".docx"]:
        text, extracted_images = dmu.read_docx_file(content, temp_folder_path)
        content_context = _header + text
        if extracted_images:
            image_file = extracted_images
    elif file_ext in [".xlsx"]:
        text = dmu.read_xlsx_file(content)
        content_context = _header + text
    elif file_ext in [".pptx"]:
        text, extracted_images = dmu.read_pptx_file(content, temp_folder_path)
        content_context = _header + text
        if extracted_images:
            image_file = extracted_images
    elif file_ext in [".pdf"] or (file_type and "pdf" in file_type):
        pdf_text, extracted_images = dmu.read_pdf_with_images(content, temp_folder_path)
        content_context = _header + json.dumps(pdf_text, ensure_ascii=False)
        if extracted_images:
            image_file = extracted_images
    elif file_type and "text" in file_type:
        content_context = _header + dmu.read_text_file(content)
    elif file_type and "json" in file_type:
        content_context = _header + json.dumps(dmu.read_json_file(content), ensure_ascii=False)
    elif file_type and "image" in file_type:
        art_critics_agent_file = support_agent["ART_CRITICS"]
        _, _, response, model_name, prompt_tokens, response_tokens = dmt.art_critics({}, {}, image_paths=[content], agent_file=art_critics_agent_file)
        content_context = _header + response
        image_file = content
    elif file_type and "audio" in file_type:
        content_context = _header + dmu.mp3_to_text(content)

    # Record the content
    content_records = {"from": content, "to":{"file_name": file_name, "file_type": file_type, "file_size": file_size, "context": content_context}}

    return content_context, content_records, image_file

# Fetch the prompt template.
# Raises a descriptive KeyError when the requested template is missing so that
# operators can immediately see *which* template name was looked up in *which*
# master file -- instead of just `KeyError: 'User Memory History'`.
# An empty/whitespace template body is treated as "template not present" so we
# also surface that case rather than returning an empty prompt silently.
def set_prompt_template(prompt_temp_cd):
    prompt_temp_mst_path = str(Path(mst_folder_path) / prompt_template_mst_file)
    prompt_temps_json = dmu.read_json_file(prompt_temp_mst_path)
    section = prompt_temps_json.get("PROMPT_TEMPLATE") or {}
    if prompt_temp_cd not in section:
        available = sorted(section.keys())
        raise KeyError(
            f"Prompt template '{prompt_temp_cd}' is missing in "
            f"{prompt_temp_mst_path} -> 'PROMPT_TEMPLATE'. "
            f"Add the template under that key (see sample_prompt_templates.json for "
            f"a reference body). Available keys ({len(available)}): {available}"
        )
    prompt_temp = section[prompt_temp_cd]
    prompt_template = ""
    if prompt_temp and str(prompt_temp).strip():
        prompt_template = prompt_template + prompt_temp + "\n"
    return prompt_template

# Get the RAG data list (in rags.json order)
def get_rag_list():
    # Use the rags.json bucket order as the basis
    rag_mst = dmu.read_json_file(rag_mst_file, mst_folder_path)
    ordered = []
    for rag_id, setting in rag_mst.items():
        bucket = setting.get("bucket", rag_id)
        if bucket not in ordered:
            ordered.append(bucket)

    # Collections that actually exist in ChromaDB
    db_client = get_chroma_client()
    collections = db_client.list_collections()
    existing = {col.name for col in collections}

    # Sort by rags.json order, keeping only those present. Append the rest alphabetically at the end.
    result = [name for name in ordered if name in existing]
    rest = sorted(existing - set(result))
    return result + rest

# Fetch the metadata list for a RAG collection (used by Knowledge Explorer)
def get_rag_collection_data(collection_name, where=None):
    """Return all metadata in a ChromaDB collection as a list of dicts.
    Pass `where` (a ChromaDB metadata-filter dict) to restrict to matching chunks only."""
    db_client = get_chroma_client()
    try:
        collection = db_client.get_collection(collection_name)
    except Exception:
        return []
    _get_kwargs = {"include": ["metadatas"]}
    if where:
        _get_kwargs["where"] = where
    response = collection.get(**_get_kwargs)
    if not response or not response["ids"]:
        return []
    data = []
    for i, cid in enumerate(response["ids"]):
        # Defensive str() — ChromaDB always returns str ids, but enforcing it
        # here means callers can safely do equality / merge / isin without
        # extra type checks regardless of which collection produced the row.
        row = {"id": str(cid)}
        row.update(response["metadatas"][i])
        data.append(row)
    return data

# Fetch the PageIndex metadata list (used by Knowledge Explorer)
def get_page_index_list():
    """Return every PageIndex dataset under pages/."""
    result = {}
    if not os.path.exists(rag_folder_pages_path):
        return result
    for folder_name in sorted(os.listdir(rag_folder_pages_path)):
        index_path = str(Path(rag_folder_pages_path) / folder_name / "_index.json")
        if os.path.exists(index_path):
            index_data = dmu.read_json_file(index_path)
            result[folder_name] = index_data.get("PAGES", [])
    return result

# Fetch RAG data
def select_rag_vector(rag_data_list, rag={}):
    buffer = 100
    rag_all = []
    rag_selected = []
    rag_context = ""

    # Current date
    current_date = datetime.now()

    # Select RAG text
    for rag_data in rag_data_list:
        rag_data["rag_name"] = rag["RAG_NAME"]
        if rag["TIMESTAMP"]=="CREATE_DATE":
            if rag_data["create_date"]:
                timestamp = datetime.strptime(dmu.convert_to_ymd(rag_data["create_date"], "%Y-%m-%d"), "%Y-%m-%d")
            else:
                timestamp = current_date
        elif rag["TIMESTAMP"]=="CURRENT_DATE" or not rag["TIMESTAMP"]:
            timestamp = current_date
        else:
            timestamp = datetime.strptime(dmu.convert_to_ymd(rag["TIMESTAMP"], "%Y-%m-%d"), "%Y-%m-%d")

        # Embed-side date setting
        timestamp_str = timestamp.strftime(rag["TIMESTAMP_STYLE"])
        days_difference = (current_date - timestamp).days
        rag_data["timestamp"] = timestamp_str
        rag_data["days_difference"] = days_difference

        # Format chunks into context via the template
        chunk_item_list = re.findall(r"\{(.*?)\}", rag["CHUNK_TEMPLATE"])
        chunk_items = {}
        for item in chunk_item_list:
            chunk_items[item] = rag_data[item]
        rag_data["chunk_context"] = rag["CHUNK_TEMPLATE"].format(**chunk_items)
        rag_data["log_format"] = rag["LOG_TEMPLATE"]

        rag_all.append(rag_data)

    # Sort by similarity
    rag_all_sorted = sorted(rag_all, key=lambda x: x["similarity_prompt"])

    # Select RAG text (up to the text cap)
    rag_context += rag["HEADER_TEMPLATE"]
    total_char = len(rag["HEADER_TEMPLATE"])
    for rag_data in rag_all_sorted:
        chunk_len = len(rag_data["chunk_context"])
        if total_char + chunk_len + buffer > rag["TEXT_LIMITS"]:
            break
        rag_selected.append(rag_data)
        rag_context += rag_data["chunk_context"]
        total_char = total_char + chunk_len + buffer

    return rag_context, rag_selected

# C-2: Build where_limitation from FILTER (precomputed since it does not depend on query_vec)
def _build_where_limitation(rag_data, exec_info, define_code={}):
    where_limitation = []
    if "FILTER" not in rag_data:
        return where_limitation
    where_limitation_conditions = []
    cond = rag_data["FILTER"]["CONDITION"]

    if "SERVICE_INFO" in cond:
        items = [{c: {"$eq": exec_info["SERVICE_INFO"]["SERVICE_ID"]}} for c in cond["SERVICE_INFO"]["ITEMS"]]
        where_limitation_conditions.append(
            items[0] if len(items) == 1 else
            {"$and" if cond["SERVICE_INFO"]["PATTERN"] == "and" else "$or": items})

    if "USER_INFO" in cond:
        items = [{c: {"$eq": exec_info["USER_INFO"]["USER_ID"]}} for c in cond["USER_INFO"]["ITEMS"]]
        where_limitation_conditions.append(
            items[0] if len(items) == 1 else
            {"$and" if cond["USER_INFO"]["PATTERN"] == "and" else "$or": items})

    if "DEFINE_CODE" in cond:
        # define_code[k] is a string or list of strings. $in for a list; $eq for a single string
        items = []
        for k, v in cond["DEFINE_CODE"]["CODES"].items():
            code_val = define_code.get(k)
            if isinstance(code_val, list):
                non_empty = [x for x in code_val if x not in ("", None)]
                if non_empty:
                    items.append({v: {"$in": non_empty}})
            elif code_val not in ("", None):
                items.append({v: {"$eq": code_val}})
        if items:
            where_limitation_conditions.append(
                items[0] if len(items) == 1 else
                {"$and" if cond["DEFINE_CODE"]["PATTERN"] == "and" else "$or": items})

    if len(where_limitation_conditions) == 1:
        where_limitation.append(where_limitation_conditions[0])
    elif len(where_limitation_conditions) > 1:
        op = "$and" if rag_data["FILTER"]["PATTERN"] == "and" else "$or"
        where_limitation.append({op: where_limitation_conditions})
    return where_limitation


# C-2: Single-query_vec ChromaDB query (the parallel-execution unit)
def _query_collection_single(collection, query_vec, result_limit, where_limitation, rag_data, meta_searches, query_seq):
    results = []

    # META_SEARCH query (date condition + similarity bonus)
    if "META_SEARCH" in rag_data:
        query_conditions_add = []
        for meta_search in meta_searches:
            if "DATE" in meta_search and "DATE" in rag_data["META_SEARCH"]["CONDITION"]:
                for date_range in meta_search["DATE"]:
                    try:
                        start_date = datetime.strptime(date_range["start"], '%Y/%m/%d').timestamp()
                        end_date = datetime.strptime(date_range["end"], '%Y/%m/%d').timestamp()
                        query_conditions_add.append({"$and": [
                            {"create_date_ts": {"$gte": start_date}},
                            {"create_date_ts": {"$lte": end_date}}
                        ]})
                    except Exception as e:
                        logger.warning("Exception: %s", e)
                        continue
        if query_conditions_add:
            where_add = query_conditions_add[0] if len(query_conditions_add) == 1 else {"$or": query_conditions_add}
            if where_limitation:
                wl = where_limitation.copy()
                wl.extend(where_add["$and"]) if "$and" in where_add else wl.append(where_add)
                where_clause = wl[0] if len(wl) == 1 else {"$and": wl}
            else:
                where_clause = where_add
            rag_data_db = collection.query(
                query_embeddings=[query_vec], n_results=result_limit,
                include=["metadatas", "embeddings", "distances"], where=where_clause)
            for i in range(len(rag_data_db["ids"])):
                for j in range(len(rag_data_db["ids"][i])):
                    v = {"id": str(rag_data_db["ids"][i][j])}
                    v |= rag_data_db["metadatas"][i][j]
                    v["vector_data_value_text"] = ast.literal_eval(v["vector_data_value_text"])
                    v["vector_data_key_text"] = rag_data_db["embeddings"][i][j].tolist()
                    v["similarity_prompt"] = round(rag_data_db["distances"][i][j], 3) * rag_data["META_SEARCH"]["BONUS"]
                    v["similarity_prompt_original"] = round(rag_data_db["distances"][i][j], 3)
                    v["query_seq"] = query_seq
                    v["query_mode"] = "(META_SEARCH:" + str(rag_data["META_SEARCH"]["BONUS"]) + ")"
                    results.append(v)

    # Standard query
    if where_limitation:
        where_clause = where_limitation[0] if len(where_limitation) == 1 else {"$and": where_limitation}
        rag_data_db = collection.query(
            query_embeddings=[query_vec], n_results=result_limit,
            include=["metadatas", "embeddings", "distances"], where=where_clause)
    else:
        rag_data_db = collection.query(
            query_embeddings=[query_vec], n_results=result_limit,
            include=["metadatas", "embeddings", "distances"])
    for i in range(len(rag_data_db["ids"])):
        for j in range(len(rag_data_db["ids"][i])):
            v = {"id": str(rag_data_db["ids"][i][j])}
            v |= rag_data_db["metadatas"][i][j]
            v["vector_data_value_text"] = ast.literal_eval(v["vector_data_value_text"])
            v["vector_data_key_text"] = rag_data_db["embeddings"][i][j].tolist()
            v["similarity_prompt"] = round(rag_data_db["distances"][i][j], 3)
            v["similarity_prompt_original"] = round(rag_data_db["distances"][i][j], 3)
            v["query_seq"] = query_seq
            v["query_mode"] = "NORMAL"
            results.append(v)
    return results


# C-2 extension: helper that runs every query_vec query for one collection together
def _process_rag_data(rag_data, query_vecs, exec_info, define_code, meta_searches, private_mode=False):
    results = []
    try:
        collection = get_chroma_client().get_collection(rag_data["DATA_NAME"])
        result_limit = min(50, collection.count())
        where_limitation = _build_where_limitation(rag_data, exec_info, define_code)
        # In Private Mode, exclude rows where private=True
        if private_mode:
            where_limitation.append({"private": {"$ne": True}})
        with ThreadPoolExecutor(max_workers=len(query_vecs)) as executor:
            futures = [
                executor.submit(
                    _query_collection_single,
                    collection, qv, result_limit, where_limitation,
                    rag_data, meta_searches, qi)
                for qi, qv in enumerate(query_vecs)
            ]
            for future in futures:
                results.extend(future.result())
    except NotFoundError:
        logger.warning(f"{rag_data['DATA_NAME']} does not exist; skipping.")
    return results

# Build context from RAG
def create_rag_context(query, query_vecs=[], rags=[], exec_info={}, meta_searches=[], define_code={}, private_mode=False):
    rag_final_context = ""
    rag_final_selected = []

    # Process per RAG dataset
    for rag in rags:
        # PageIndex type: skip vector search and go to its dedicated path
        if rag.get("RETRIEVER") == "PageIndex":
            rag_context, rag_selected = select_rag_page_index(query, rag, exec_info)
            rag_final_context += rag_context
            rag_final_selected += rag_selected
            continue

        # AgentSearch type: invoke another agent (incl. self), bounded by a
        # shared recursion counter seeded by DigiMatsuExecute.
        if rag.get("RETRIEVER") == "AgentSearch":
            rag_context, rag_selected = select_rag_agent_search(query, rag, exec_info)
            rag_final_context += rag_context
            rag_final_selected += rag_selected
            continue

        # FunctionSearch type: invoke a registered tool function.
        if rag.get("RETRIEVER") == "FunctionSearch":
            rag_context, rag_selected = select_rag_function_search(query, rag, exec_info)
            rag_final_context += rag_context
            rag_final_selected += rag_selected
            continue

        # Parallel processing for DB-type rag_data (C-2 extension: also parallelizes per collection)
        db_rag_data_list = [rd for rd in rag["DATA"] if rd.get("DATA_TYPE") == "DB"]
        rag_data_list = []
        if db_rag_data_list:
            with ThreadPoolExecutor(max_workers=len(db_rag_data_list)) as executor:
                futures = [
                    executor.submit(_process_rag_data, rd, query_vecs, exec_info, define_code, meta_searches, private_mode)
                    for rd in db_rag_data_list
                ]
                for future in futures:
                    rag_data_list.extend(future.result())

        # In rag_data_list, dedupe by id keeping the one with the closest "query similarity"
        filtered_data = {}
        for rag_data in rag_data_list:
            rag_data_id = rag_data["id"]
            if rag_data_id not in filtered_data:
                filtered_data[rag_data_id] = rag_data
            elif rag_data["similarity_prompt"] < filtered_data[rag_data_id]["similarity_prompt"]:
                filtered_data[rag_data_id] = rag_data
        rag_data_list = list(filtered_data.values())

        # Select the RAG data
        if rag.get("RETRIEVER") == "Vector":
            rag_context, rag_selected = select_rag_vector(rag_data_list, rag)
            rag_final_context += rag_context
            rag_final_selected += rag_selected

    if rag_final_context:
        rag_final_context += "----\nこれらの情報を踏まえて、次の質問に日本語で回答してください。\n----\n"  # Final RAG prompt suffix (kept JP: sent to the LLM)

    return rag_final_context, rag_final_selected

# Evaluate similarity between the response and the RAG chunks
def get_knowledge_reference(response_vec, rag_selected, logic="Cosine"):
    rag_ref = []

    # Evaluate similarity per chunk
    for rag_data in rag_selected:
        # PageIndex-derived string entries are appended to the log as-is and skipped
        if isinstance(rag_data, str):
            rag_ref.append(rag_data)
            continue
        rag_data["value_text_short"] = rag_data["value_text"][:50]  # truncate to 50 chars
        similarity_response = dmu.calculate_similarity_vec(response_vec, rag_data["vector_data_value_text"], logic)
        rag_data["similarity_response"] = round(similarity_response,3)

        # Apply the meta-search correction
        if rag_data["query_mode"] != "NORMAL":
            key, value = rag_data["query_mode"].strip("()").split(":")
            if key == "META_SEARCH":
                rag_data["similarity_prompt"] = round(rag_data["similarity_prompt"]/float(value),3)
#                rag_data["similarity_response"] = round(similarity_response*float(value),3)

        # Log format for the UI
        chunk_item_list = re.findall(r"\{(.*?)\}", rag_data["log_format"])
        chunk_items = {}
        for item in chunk_item_list:
            chunk_items[item] = rag_data[item]
        rag_data["log"] = rag_data["log_format"].format(**chunk_items)

        # Dataset kept for the record
        rag_ref.append(rag_data["log"])
    return rag_ref

# Conversation-memory reference info
def get_memory_reference(memory_selected, memory_similarity=False, response_vec=[], logic="Cosine"):
    memory_ref = []

    for memory_data in memory_selected:
        seq = memory_data["seq"]
        sub_seq = memory_data["sub_seq"]
        type = memory_data["type"]
        timestamp = memory_data["timestamp"]
        text = memory_data["text"]

        # Log format for the UI
        memory_ref_log = f"Chat history at {timestamp}: {seq}_{sub_seq}_{type} \"{text[:50]}\"<br>"
        if memory_similarity and response_vec:
            similarity_prompt = round(memory_data["similarity_prompt"],3)
            similarity_response = round(dmu.calculate_similarity_vec(response_vec, memory_data["vec_text"], logic),3)
            memory_ref_log = f"Chat history at {timestamp}: {seq}_{sub_seq}_{type} [query similarity: {round(similarity_prompt,3)}, response similarity: {round(similarity_response,3)}] {text[:50]}<br>"

        # Dataset kept for the record
        memory_ref.append(
            {
                "log": memory_ref_log
            }
        )
    return memory_ref

# Build RAG chunk data from CSV (utf-8)
def get_chunk_csv(bucket, file_path, file_name, field_items, title_items, key_text_items, value_text_items, category_items=[]):
    rag_data = []

    csv_path = Path(file_path) / file_name
    if not csv_path.exists():
        logger.warning(f"CSV file is missing: {file_name}")
        return rag_data

    with open(csv_path, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile, fieldnames=field_items)
        next(reader, None)

        for i, row in enumerate(reader):
            data_matched = True

            # Filtering
            if category_items:
                for cond in category_items:
                    for key, values in cond.items():
                        cell_value = row.get(key, "").strip().lower()
                        target_values = [v.lower() for v in values]
                        if cell_value not in target_values:
                            data_matched = False

            # Create a chunk when the conditions match
            if data_matched:
                rag_chunk = {}
                rag_chunk["id"] = bucket+"-"+str(i+1)
                rag_chunk["bucket"] = bucket

                title = ""
                for title_item in title_items:
                    if title_item in row:
                        title += row[title_item]
                    else:
                        title += title_item
                rag_chunk["title"] = title

                key_text = ""
                for key_text_item in key_text_items:
                    if key_text_item in row:
                        key_text += row[key_text_item]
                    else:
                        key_text += key_text_item
                rag_chunk["key_text"] = key_text

                value_text = ""
                for value_text_item in value_text_items:
                    if value_text_item in row:
                        value_text += row[value_text_item]
                    else:
                        value_text += value_text_item
                rag_chunk["value_text"] = value_text

                for field_item in field_items:
                    rag_chunk[field_item] = row[field_item]

                create_date = datetime.now().strftime("%Y-%m-%d")
                if "create_date" in row:
                    try:
                        parsed_date = datetime.strptime(row["create_date"], '%Y/%m/%d')
                        create_date = parsed_date.strftime('%Y-%m-%d')
                    except ValueError:
                        create_date = datetime.now().strftime("%Y-%m-%d")
                rag_chunk["create_date"] = create_date

                rag_data.append(rag_chunk)
    return rag_data

# Build RAG chunk data from a Notion database
def get_chunk_notion(bucket, db_name, item_dict, chk_dict=None, date_dict=None, category_dict=None):
    rag_data = []
    
    # Resolve the Notion DB ID
    notion_db_mst_file_path = str(Path(mst_folder_path) / notion_db_mst_file)
    notion_db_mst = dmu.read_json_file(notion_db_mst_file_path)
    db_id = notion_db_mst[db_name]

    # Fetch RAG-target pages
    pages = dmn.get_pages_done(db_id, chk_dict, date_dict, category_dict)

    # Convert to RAG-data form. Force str() on every Notion-derived id so
    # downstream lookups (chat-log id_color_map, scatter merge/isin) stay
    # consistent with ChromaDB's string-only id contract.
    page_ids = [str(page['id']) for page in pages]
    for page_id in page_ids:
        if item_dict is not None:
            page_items = {}
            page_items.update({'id': page_id})
            page_items.update({'bucket': bucket})
            skip_page = False
            for key, value in item_dict.items():
                if isinstance(value, dict):
                    for k, v in value.items():
                        item_val = dmn.get_notion_item_by_id(pages, page_id, k, v)
                        if item_val is None:
                            logger.warning(f"[SKIP] Notion page {page_id}: property \"{k}\" ({v}) is unset; skipping")
                            skip_page = True
                            break
                        page_items[key] = item_val
                    if skip_page:
                        break
                elif isinstance(value, list):
                    page_item_text = ""
                    for item in value:
                        i = 0
                        if i != 0:
                            page_item_text += "\n"
                        if isinstance(item, dict):
                            for k, v in item.items():
                                page_item_text += dmn.get_notion_item_by_id(pages, page_id, k, v) or ""
                        else:
                            page_item_text += item
                    page_items[key] = page_item_text
                else:
                    page_items[key] = value

            if skip_page:
                continue

            if "create_date" not in page_items:
                page_items["create_date"] = datetime.now().strftime("%Y-%m-%d")

            rag_data.append(page_items)
    return rag_data

# PageIndex: build chunks from a Notion database (with body content)
def get_chunk_notion_pageindex(bucket, db_name, item_dict, chk_dict=None, date_dict=None, category_dict=None):
    rag_data = []

    # Resolve the Notion DB ID
    notion_db_mst_file_path = str(Path(mst_folder_path) / notion_db_mst_file)
    notion_db_mst = dmu.read_json_file(notion_db_mst_file_path)
    db_id = notion_db_mst[db_name]

    # Fetch RAG-target pages
    pages = dmn.get_pages_done(db_id, chk_dict, date_dict, category_dict)

    for page in pages:
        notion_page_id = str(page["id"])
        if item_dict is None:
            continue

        page_items = {"notion_page_id": notion_page_id, "bucket": bucket}
        skip_page = False
        for key, value in item_dict.items():
            if isinstance(value, dict):
                for k, v in value.items():
                    item_val = dmn.get_notion_item_by_id(pages, notion_page_id, k, v)
                    if item_val is None:
                        logger.warning(f"[SKIP] Notion page {notion_page_id}: property \"{k}\" ({v}) is unset; skipping")
                        skip_page = True
                        break
                    page_items[key] = item_val
                if skip_page:
                    break
            elif isinstance(value, list):
                page_item_text = ""
                for item in value:
                    if isinstance(item, dict):
                        for k, v in item.items():
                            page_item_text += dmn.get_notion_item_by_id(pages, notion_page_id, k, v) or ""
                    else:
                        page_item_text += item
                page_items[key] = page_item_text
            else:
                page_items[key] = value
        if skip_page:
            continue

        # Fetch the page body and store as `body`
        try:
            page_items["body"] = dmn.get_page_body_text(notion_page_id)
        except Exception as e:
            logger.warning(f"Failed to fetch Notion page body (page={notion_page_id}): {e}")
            page_items["body"] = ""

        if "create_date" not in page_items:
            page_items["create_date"] = datetime.now().strftime("%Y-%m-%d")

        rag_data.append(page_items)
    return rag_data


_PAGEINDEX_BODY_EXTS = {".txt", ".md"}


# If the body cell value is a .txt/.md filename under source_dir, return that file's contents.
# Otherwise return the cell value verbatim.
# To mitigate path traversal, only bare filenames are allowed (anything containing a separator or '..' is treated as inline).
def _resolve_pageindex_body(source_dir, value):
    if not isinstance(value, str):
        return ""
    candidate = value.strip()
    if not candidate:
        return ""
    if any(sep in candidate for sep in ("/", "\\", "..")):
        return value
    if Path(candidate).suffix.lower() not in _PAGEINDEX_BODY_EXTS:
        return value
    file_path = Path(source_dir) / candidate
    if not file_path.is_file():
        return value
    try:
        return file_path.read_text(encoding="utf-8")
    except Exception as e:
        logger.warning(f"Failed to read body file ({file_path}): {e}")
        return value


# Build PageIndex chunks from an Excel sheet (one row = one page)
def get_chunk_excel_pageindex(bucket, source_dir, source_file, sheet, item_dict):
    rag_data = []
    src_path = Path(source_dir) / source_file
    if not src_path.exists():
        logger.warning(f"Excel source not found: {src_path}")
        return rag_data

    try:
        df = pd.read_excel(str(src_path), sheet_name=sheet, dtype=str).fillna("")
    except Exception as e:
        logger.error(f"Excel load failed ({src_path}, sheet={sheet}): {e}")
        return rag_data

    if not isinstance(item_dict, dict):
        logger.warning(f"Invalid item_dict: {item_dict}")
        return rag_data

    for _, row in df.iterrows():
        page_items = {"bucket": bucket}
        for key, col in item_dict.items():
            if not col:
                page_items[key] = ""
                continue
            val = row.get(col, "")
            page_items[key] = val.strip() if isinstance(val, str) else val

        page_id = str(page_items.get("id", "")).strip()
        if not page_id:
            logger.warning(f"[SKIP] Excel row: ID column \"{item_dict.get('id')}\" is empty; skipping")
            continue
        page_items["id"] = page_id

        page_items["body"] = _resolve_pageindex_body(source_dir, page_items.get("body", ""))

        tags = page_items.get("tags", "")
        if isinstance(tags, str):
            page_items["tags"] = [t.strip() for t in re.split(r"[,|]", tags) if t.strip()]

        if not page_items.get("create_date"):
            page_items["create_date"] = datetime.now().strftime("%Y-%m-%d")

        rag_data.append(page_items)
    return rag_data


# Dynamically compute sort_order from the id ("1-0" -> 100, "1-2" -> 102, "1-2-3" -> 10203)
def _derive_sort_order(page_id):
    parts = str(page_id).split("-")
    sort_order = 0
    for i, p in enumerate(parts):
        try:
            num = int(p)
        except ValueError:
            num = 0
        weight = 100 ** (len(parts) - 1 - i)
        sort_order += num * weight
    return sort_order


# Save PageIndex RAG data (from Notion; duplicates by id overwrite)
def save_rag_pageindex(bucket, rag_data):
    pages_dir = str(Path(rag_folder_pages_path) / bucket)
    os.makedirs(pages_dir, exist_ok=True)

    index_path = str(Path(pages_dir) / "_index.json")
    index_data = dmu.read_json_file(index_path) if os.path.exists(index_path) else {}
    # Coerce existing-JSON keys to str so a legacy index that happens to
    # carry integer ids merges cleanly with the freshly-ingested str ids.
    pages_map = {str(p["id"]): {**p, "id": str(p["id"])}
                  for p in index_data.get("PAGES", []) if "id" in p}

    book_title = None
    processed = []

    for chunk in rag_data:
        page_id = str(chunk.get("id", "")).strip()
        if not page_id:
            logger.warning(f"[SKIP] cannot save to PageIndex because id is unset (notion={chunk.get('notion_page_id')})")
            continue

        if book_title is None and chunk.get("book"):
            book_title = chunk["book"]

        title = chunk.get("title", page_id)
        body = chunk.get("body", "")
        md_content = f"# {title}\n\n{body}" if body else f"# {title}\n"
        md_path = str(Path(pages_dir) / f"{page_id}.md")
        dmu.save_text_file(md_content, md_path)

        tags = chunk.get("tags", [])
        if isinstance(tags, str):
            tags = [t.strip() for t in tags.split(",") if t.strip()]

        entry = {
            "id": page_id,
            "title": title,
            "timestamp": chunk.get("create_date", ""),
            "summary": chunk.get("summary", ""),
            "tags": tags,
            "category": chunk.get("category", ""),
            "sort_order": _derive_sort_order(page_id),
        }
        pages_map[page_id] = entry
        processed.append({"id": page_id, "notion_page_id": chunk.get("notion_page_id")})

    # Set BOOK info (first matching book value, or keep existing, or fall back to bucket)
    if book_title:
        index_data["BOOK"] = {"title": book_title}
    elif "BOOK" not in index_data:
        index_data["BOOK"] = {"title": bucket}

    # Sort by sort_order before storing
    sorted_pages = sorted(pages_map.values(), key=lambda x: x.get("sort_order", 0))
    index_data["PAGES"] = sorted_pages

    dmu.save_json_file(index_data, index_path)
    logger.info(f"{bucket}: saved {len(processed)} pages -> {pages_dir}")
    return processed


# Export PageIndex RAG as a ZIP bundle of Excel + individual .md files (returns bytes).
# The output Excel format is compatible with the input='excel' pageindex importer (body column stores the file name).
def export_pageindex_as_excel_bundle(bucket):
    import io
    import zipfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    pages_dir = Path(rag_folder_pages_path) / bucket
    index_path = pages_dir / "_index.json"
    if not index_path.exists():
        return None

    index_data = dmu.read_json_file(str(index_path))
    book_title = index_data.get("BOOK", {}).get("title", bucket)
    pages = sorted(index_data.get("PAGES", []), key=lambda p: p.get("sort_order", 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "pages"
    headers = ["ブック名", "ID", "タイトル", "サマリー", "タグ", "カテゴリ", "本文"]  # Excel column headers (kept JP: match the importer schema)
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for p in pages:
        page_id = str(p.get("id", ""))
        tags_val = p.get("tags", [])
        if isinstance(tags_val, list):
            tags_val = ",".join(tags_val)
        body_ref = f"{page_id}.md" if (pages_dir / f"{page_id}.md").exists() else ""
        ws.append([
            book_title, page_id,
            p.get("title", ""), p.get("summary", ""),
            tags_val, p.get("category", ""), body_ref,
        ])

    widths = [22, 8, 36, 50, 28, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{bucket}/{bucket}.xlsx", xlsx_buf.getvalue())
        for p in pages:
            page_id = str(p.get("id", ""))
            md_path = pages_dir / f"{page_id}.md"
            if md_path.exists():
                zf.writestr(f"{bucket}/{page_id}.md", md_path.read_bytes())
    return zip_buf.getvalue()


# Edit RAG chunk data (ChromaDB)
def save_rag_chunk_db(rag_id, rag_data):
    db_client = get_chroma_client()
    collection = db_client.get_or_create_collection(name=rag_id, metadata={"hnsw:space": "cosine"})
    response = collection.get(include=["metadatas"])
    # All keys in `existing_map` and every ChromaDB id sent below are str()'d.
    # ChromaDB itself stores ids as strings; the defensive coercion guards
    # against numeric ids leaking in from upstream ingestion paths (Excel
    # numeric cell, JSON int) so downstream dict/isin/merge lookups don't
    # break on a "5" vs 5 mismatch.
    existing_map = {}
    if response and "ids" in response:
        for i, cid in enumerate(response["ids"]):
            existing_map[str(cid)] = response["metadatas"][i]
    cnt_add = 0
    cnt_extent = 0

    # Sort RAG data by create_date descending
    if rag_data and "create_date" in rag_data[0]:
        rag_data = sorted(rag_data, key=lambda x: x["create_date"], reverse=True)

    # Bucket chunks by what changed
    chunks_need_embed = []  # New, or title/key_text/value_text changed -> needs embedding
    chunks_meta_only  = []  # Only other fields changed -> metadata-only overwrite

    skip_keys_for_compare = {"id", "title", "key_text", "value_text", "vector_data_value_text"}

    for rag_chunk in rag_data:
        rag_chunk["create_date_ts"] = datetime.strptime(rag_chunk["create_date"], "%Y-%m-%d").timestamp()
        if "private" not in rag_chunk:
            rag_chunk["private"] = False
        if not rag_chunk.get("value_text"):
            continue

        # Normalise the chunk id to str() so all comparisons against
        # existing_map (also string-keyed) work uniformly.
        rag_chunk["id"] = str(rag_chunk["id"])
        chunk_id = rag_chunk["id"]
        if chunk_id not in existing_map:
            chunks_need_embed.append(rag_chunk)
        else:
            existing = existing_map[chunk_id]
            text_changed = (
                rag_chunk["title"]      != existing.get("title") or
                rag_chunk["key_text"]   != existing.get("key_text") or
                rag_chunk["value_text"] != existing.get("value_text")
            )
            if text_changed:
                chunks_need_embed.append(rag_chunk)
            else:
                meta_changed = any(
                    str(rag_chunk.get(k)) != str(existing.get(k))
                    for k in rag_chunk
                    if k not in skip_keys_for_compare
                )
                if meta_changed:
                    chunks_meta_only.append(rag_chunk)
                else:
                    logger.info(f"{rag_chunk['title']} already exists in the knowledge DB (no changes).")
                    cnt_extent += 1

    # Process chunks needing embedding in a single API call (send key_text + value_text together)
    if chunks_need_embed:
        key_texts = [c["key_text"].replace("\n", "") for c in chunks_need_embed]
        val_texts = [c["value_text"].replace("\n", "") for c in chunks_need_embed]
        all_vecs  = dmu.embed_texts_batch(key_texts + val_texts)
        n = len(chunks_need_embed)
        key_vecs = all_vecs[:n]
        val_vecs = all_vecs[n:]

        for i, rag_chunk in enumerate(chunks_need_embed):
            chunk_id = str(rag_chunk["id"])
            del rag_chunk["id"]
            rag_chunk["vector_data_value_text"] = str(val_vecs[i])
            if chunk_id in existing_map:
                collection.delete(ids=[chunk_id])
                logger.info(f"Updated {rag_chunk['title']} in the knowledge DB.")
            else:
                logger.info(f"Added {rag_chunk['title']} to the knowledge DB.")
            collection.add(ids=[chunk_id], embeddings=[key_vecs[i]], metadatas=rag_chunk)
            cnt_add += 1

    # Overwrite metadata-only chunks without re-embedding
    for rag_chunk in chunks_meta_only:
        chunk_id = str(rag_chunk["id"])
        del rag_chunk["id"]
        rag_chunk["vector_data_value_text"] = existing_map[chunk_id].get("vector_data_value_text", "")
        collection.update(ids=[chunk_id], metadatas=rag_chunk)
        logger.info(f"Updated metadata for {rag_chunk['title']} in the knowledge DB (no re-embedding).")
        cnt_add += 1

    return cnt_add, cnt_extent

# Migration helper: bulk-add the private flag to existing RAG data
def migrate_add_private_flag():
    db_client = get_chroma_client()
    collections = db_client.list_collections()
    total = 0
    for col in collections:
        col_name = col.name if hasattr(col, 'name') else str(col)
        collection = db_client.get_collection(col_name)
        response = collection.get(include=["metadatas"])
        if not response or not response["ids"]:
            continue
        ids_to_update = []
        metas_to_update = []
        for i, cid in enumerate(response["ids"]):
            meta = response["metadatas"][i]
            if "private" not in meta:
                meta["private"] = False
                ids_to_update.append(cid)
                metas_to_update.append(meta)
        if ids_to_update:
            collection.update(ids=ids_to_update, metadatas=metas_to_update)
            total += len(ids_to_update)
            logger.info(f"{col_name}: applied private=False to {len(ids_to_update)} rows")
    logger.info(f"Migration done: {total} rows total")
    return total


# Save PageIndex RAG data (.md + _index.json)
def save_rag_pages(rag_id, rag_data):
    pages_dir = str(Path(rag_folder_pages_path) / rag_id)
    os.makedirs(pages_dir, exist_ok=True)

    index_pages = []
    cnt = 0
    for chunk in rag_data:
        # `.strip()` is unsafe on non-str sources; coerce first so a numeric
        # page_id from upstream doesn't crash the loop.
        page_id = str(chunk.get("page_id", "") or "").strip()
        if not page_id:
            continue

        # Save the page body as .md
        body = chunk.get("body", chunk.get("value_text", ""))
        title = chunk.get("title", page_id)
        md_content = f"# {title}\n\n{body}"
        md_path = str(Path(pages_dir) / f"{page_id}.md")
        dmu.save_text_file(md_content, md_path)

        # Create the index entry
        tags_raw = chunk.get("tags", "")
        tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if isinstance(tags_raw, str) else tags_raw
        index_entry = {
            "id": page_id,
            "title": title,
            "summary": chunk.get("summary", ""),
            "tags": tags,
            "category": chunk.get("category", ""),
        }
        sort_order = chunk.get("sort_order")
        if sort_order is not None:
            index_entry["sort_order"] = sort_order
        index_pages.append(index_entry)
        cnt += 1

    # Sort by sort_order if present
    if index_pages and "sort_order" in index_pages[0]:
        index_pages.sort(key=lambda x: x.get("sort_order", 0))

    # Save _index.json
    index_data = {"PAGES": index_pages}
    index_path = str(Path(pages_dir) / "_index.json")
    dmu.save_json_file(index_data, index_path)

    logger.info(f"{rag_id}: saved {cnt} pages -> {pages_dir}")
    return cnt, len(index_pages)


# Augment a user-provided LOG_TEMPLATE so the resulting log string carries the
# fields `analytics_knowledge` requires (similarity_Q, similarity_A, DB, ID,
# QUERY_SEQ, QUERY_MODE). Only missing keys are appended — Vector templates
# already include these, so this is a no-op for them. Used by the non-Vector
# retrievers (PageIndex / AgentSearch / FunctionSearch) so their chunks show
# up in Analytics Result - Knowledge Utility too.
def _augment_log_template(template):
    appendix = []
    if "similarity_Q" not in template:
        appendix.append("'similarity_Q': {similarity_prompt}")
    if "similarity_A" not in template:
        appendix.append("'similarity_A': {similarity_response}")
    if "'DB'" not in template:
        appendix.append("'DB': '{bucket}'")
    if "'ID'" not in template:
        appendix.append("'ID': '{id}'")
    if "QUERY_SEQ" not in template:
        appendix.append("'QUERY_SEQ': '{query_seq}'")
    if "QUERY_MODE" not in template:
        appendix.append("'QUERY_MODE': '{query_mode}'")
    if "'title'" not in template:
        appendix.append("'title': '{title}'")
    if not appendix:
        return template
    sep = ", " if template.strip() else ""
    return template + sep + ", ".join(appendix)


# Build a "Title > Title > ..." breadcrumb from a dash-separated PageIndex id.
# E.g. id="1-1-1" → look up pages_map for "1", "1-1", "1-1-1" and join their
# titles. Missing ancestors (ids that aren't standalone pages) are skipped,
# so for fully-flat indexes the breadcrumb degenerates to just the self title.
def _build_page_breadcrumb(page_id, pages_map):
    segs = str(page_id).split("-")
    titles = []
    for _i in range(1, len(segs) + 1):
        _anc = "-".join(segs[:_i])
        _meta = pages_map.get(_anc)
        if _meta and _meta.get("title"):
            titles.append(str(_meta["title"]).strip())
    return " > ".join(titles)


# Build context from an AgentSearch RAG.
# Runs another DigiMatsu agent (incl. self) via DigiMatsuExecute_Practice and
# feeds the response back as RAG context. Recursion is capped by a shared
# `_AGENT_SEARCH_STATE` dict that DigiMatsuExecute seeds into exec_info, and
# every nested call propagates the same dict via in_execution so the counter
# is honored across the whole request.
#
# DATA[0] schema:
#   AGENT_FILE       — required, target agent JSON
#   MAX_CALLS        — per-block override of the agent root AGENT_SEARCH_MAX_CALLS
#   EXECUTION        — Execute_Practice in_execution shape; safe defaults fill the rest
#   OVERWRITE_ITEMS  — overrides for target agent (HABIT/PERSONALITY/...) — same shape as overwrite_items
#   ADD_KNOWLEDGE    — extra RAG entries injected into the child's KNOWLEDGE (BOOK injection style)
#   SITUATION        — TIME/SITUATION dict for the child
def select_rag_agent_search(query, rag, exec_info):
    import DigiM_Execute as _dme_as  # late import to dodge the import cycle
    rag_context = ""
    rag_selected = []

    service_info = exec_info.get("SERVICE_INFO", {})
    user_info = exec_info.get("USER_INFO", {})
    session_id = exec_info.get("_SESSION_ID", "")
    session_name = exec_info.get("_SESSION_NAME", "")

    # Shared cap dict — DigiMatsuExecute seeds this from in_execution before
    # calling create_rag_context. Default = 3 if a caller bypassed the seed.
    state = exec_info.setdefault("_AGENT_SEARCH_STATE", {"calls": 0, "max": 3})

    _SAFE_EXEC_DEFAULTS = {
        "MEMORY_USE": False,
        "MEMORY_SAVE": False,        # child input/output NOT persisted to chat_memory
        "MEMORY_SIMILARITY": False,
        "SAVE_DIGEST": False,
        "STREAM_MODE": False,
        "PRIVATE_MODE": True,
        "MAGIC_WORD_USE": False,
        "META_SEARCH": False,
        "RAG_QUERY_GENE": False,
        "WEB_SEARCH": False,
        "WEB_SEARCH_ENGINE": "",
        "THINKING_MODE": False,
        "USER_MEMORY_LAYERS": [],
        "INSERT_CITATIONS": True,
    }

    for rd in rag.get("DATA", []):
        if rd.get("DATA_TYPE") != "AGENT_SEARCH":
            continue
        target_agent_file = rd.get("AGENT_FILE")
        if not target_agent_file:
            logger.warning(f"AgentSearch missing AGENT_FILE for RAG={rag.get('RAG_NAME')}")
            continue

        # Per-block MAX_CALLS upper-bounds (always raises the parent's cap, never lowers)
        block_max = rd.get("MAX_CALLS")
        if block_max is not None:
            try:
                state["max"] = max(int(state.get("max", 3)), int(block_max))
            except (TypeError, ValueError):
                pass

        if state.get("calls", 0) >= state.get("max", 3):
            logger.warning(
                f"AgentSearch cap reached (max={state.get('max')}); "
                f"skipping {rag.get('RAG_NAME')} → {target_agent_file}"
            )
            continue

        state["calls"] = state.get("calls", 0) + 1
        current_idx = state["calls"]

        # Compose child execution
        child_exec = dict(_SAFE_EXEC_DEFAULTS)
        for _k, _v in (rd.get("EXECUTION") or {}).items():
            child_exec[_k] = _v
        # The shared counter dict is passed BY REFERENCE so the child's
        # nested AgentSearches keep incrementing the same value.
        child_exec["_AGENT_SEARCH_STATE"] = state
        # Parent already owns the session lock; child must not try to reclaim it.
        child_exec["_PRE_LOCKED"] = True

        overwrite_items = rd.get("OVERWRITE_ITEMS") or {}
        add_knowledge = rd.get("ADD_KNOWLEDGE") or []
        situation = rd.get("SITUATION") or {"TIME": "", "SITUATION": ""}

        # Run the child Practice and concat its streamed chunks
        response_text = ""
        ts_begin = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            for _y in _dme_as.DigiMatsuExecute_Practice(
                service_info, user_info, session_id, session_name,
                target_agent_file, query, [], situation,
                overwrite_items, add_knowledge, child_exec,
            ):
                if isinstance(_y, tuple) and len(_y) >= 3:
                    _chunk = _y[2]
                    if _chunk and not str(_chunk).startswith("[STATUS]"):
                        response_text += str(_chunk)
        except Exception as e:
            logger.warning(f"AgentSearch execution error ({target_agent_file}): {e}")
            response_text = f"[AgentSearch error] {type(e).__name__}: {e}"

        # Look up the target's display name for templates / logs
        try:
            target_agent_data = dmu.read_json_file(target_agent_file, agent_folder_path) or {}
            agent_name = (target_agent_data.get("NAME")
                          or target_agent_data.get("DISPLAY_NAME") or target_agent_file)
        except Exception:
            agent_name = target_agent_file

        chunk_template = rag.get(
            "CHUNK_TEMPLATE",
            "[{rag_name} / {agent_name}]\nQ: {query}\nA: {response}\n\n",
        )
        try:
            chunk_text = chunk_template.format(
                rag_name=rag.get("RAG_NAME", ""),
                agent_name=agent_name,
                agent_file=target_agent_file,
                query=query,
                response=response_text,
                response_tokens=len(response_text),
            )
        except (KeyError, IndexError):
            chunk_text = (f"[{rag.get('RAG_NAME', '')} / {agent_name}]\n"
                          f"Q: {query}\nA: {response_text}\n\n")
        rag_context += chunk_text

        log_template = rag.get(
            "LOG_TEMPLATE",
            "'rag':'{rag_name}', 'agent':'{agent_name}', 'tokens':'{response_tokens}'",
        )
        # Augment so Analytics Result - Knowledge Utility can score this chunk.
        log_template = _augment_log_template(log_template)

        # Embed the child response and the query for the similarity_Q signal.
        # If embedding fails we degrade gracefully: emit a legacy string log
        # entry (no analytics for this chunk) instead of crashing the turn.
        chunk_vec, sim_Q = None, 0.0
        try:
            _qv = dmu.embed_text(query)
            _cv = dmu.embed_text(response_text) if response_text else None
            if _qv and _cv:
                chunk_vec = _cv
                sim_Q = round(dmu.calculate_similarity_vec(_qv, _cv, "Cosine"), 3)
        except Exception as _e:
            logger.warning(f"AgentSearch chunk embedding failed ({target_agent_file}): {_e}")

        if chunk_vec is not None:
            rag_selected.append({
                "rag_name":  rag.get("RAG_NAME", ""),
                "rag":       rag.get("RAG_NAME", ""),
                "bucket":    "AgentSearch",
                "DB":        "AgentSearch",
                "id":        target_agent_file,
                "ID":        target_agent_file,
                "title":     agent_name,
                "agent_name": agent_name,
                "agent_file": target_agent_file,
                "query":     query,
                "response":  response_text,
                "response_tokens": len(response_text),
                "category":  agent_name,
                "value_text": response_text,
                "vector_data_value_text": chunk_vec,
                "similarity_prompt": sim_Q,
                "query_mode": "NORMAL",
                "QUERY_MODE": "NORMAL",
                "query_seq":  "0",
                "QUERY_SEQ":  "0",
                "log_format": log_template,
            })
        else:
            # Fallback: render the template directly (no similarity → no analytics row)
            try:
                _legacy = rag.get("LOG_TEMPLATE",
                    "'rag':'{rag_name}', 'agent':'{agent_name}', 'tokens':'{response_tokens}'").format(
                        rag_name=rag.get("RAG_NAME", ""),
                        agent_name=agent_name,
                        agent_file=target_agent_file,
                        response_tokens=len(response_text))
            except (KeyError, IndexError):
                _legacy = f"rag={rag.get('RAG_NAME', '')}, agent={agent_name}"
            rag_selected.append(_legacy)

        # Stash the full detail so the parent turn can persist it under
        # `prompt.agent_search` (similar to thinking_log / web_search_log).
        exec_info.setdefault("_AGENT_SEARCH_LOG", []).append({
            "call_idx": current_idx,
            "rag_name": rag.get("RAG_NAME", ""),
            "agent_file": target_agent_file,
            "agent_name": agent_name,
            "query": query,
            "response": response_text,
            "response_tokens": len(response_text),
            "timestamp": ts_begin,
        })

    if rag_context:
        header = rag.get("HEADER_TEMPLATE", "")
        rag_context = header + rag_context
        text_limits = rag.get("TEXT_LIMITS", 6000)
        if len(rag_context) > text_limits:
            rag_context = rag_context[:text_limits]
    return rag_context, rag_selected


# Build context from a FunctionSearch RAG.
# Invokes a registered tool function (see DigiM_ToolRegistry) and feeds its
# return value back as RAG context.
#
# DATA[0] schema:
#   FUNCTION_NAME    — required, name in the tool registry
#   ARGS_TEMPLATE    — string template formatted against {query}. Default = "{query}".
def select_rag_function_search(query, rag, exec_info):
    import inspect as _inspect_fs
    # `call_function_by_name` lives in DigiM_Tool (the dispatcher); the registry
    # only owns the lookup table. Late import to dodge any cycle on first load.
    import DigiM_Tool as _dmt_fs
    rag_context = ""
    rag_selected = []

    service_info = exec_info.get("SERVICE_INFO", {})
    user_info = exec_info.get("USER_INFO", {})
    session_id = exec_info.get("_SESSION_ID", "")
    session_name = exec_info.get("_SESSION_NAME", "")
    parent_agent_file = exec_info.get("_AGENT_FILE", "")

    for rd in rag.get("DATA", []):
        if rd.get("DATA_TYPE") != "FUNCTION_SEARCH":
            continue
        func_name = rd.get("FUNCTION_NAME")
        if not func_name:
            logger.warning(f"FunctionSearch missing FUNCTION_NAME for RAG={rag.get('RAG_NAME')}")
            continue
        args_template = rd.get("ARGS_TEMPLATE", "{query}")
        try:
            input_str = args_template.format(query=query)
        except (KeyError, IndexError):
            input_str = query

        response_text = ""
        ts_begin = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            result = _dmt_fs.call_function_by_name(
                service_info, user_info, func_name,
                session_id, session_name, parent_agent_file,
                input_str, [], {},
            )
            if _inspect_fs.isgenerator(result):
                for chunk_pack in result:
                    if isinstance(chunk_pack, (tuple, list)) and len(chunk_pack) >= 3:
                        _chunk = chunk_pack[2]
                        if _chunk and not str(_chunk).startswith("[STATUS]"):
                            response_text += str(_chunk)
            elif isinstance(result, (tuple, list)) and len(result) > 2:
                response_text = str(result[2]) if result[2] is not None else ""
            else:
                response_text = str(result) if result is not None else ""
        except Exception as e:
            logger.warning(f"FunctionSearch execution error ({func_name}): {e}")
            response_text = f"[FunctionSearch error] {type(e).__name__}: {e}"

        chunk_template = rag.get(
            "CHUNK_TEMPLATE",
            "[{rag_name} / {function_name}]\n{response}\n\n",
        )
        try:
            chunk_text = chunk_template.format(
                rag_name=rag.get("RAG_NAME", ""),
                function_name=func_name,
                query=query,
                args=input_str,
                response=response_text,
            )
        except (KeyError, IndexError):
            chunk_text = f"[{rag.get('RAG_NAME', '')} / {func_name}]\n{response_text}\n\n"
        rag_context += chunk_text

        log_template = rag.get(
            "LOG_TEMPLATE",
            "'rag':'{rag_name}', 'function':'{function_name}'",
        )
        # Augment so Analytics Result - Knowledge Utility can score this chunk.
        log_template = _augment_log_template(log_template)

        chunk_vec, sim_Q = None, 0.0
        try:
            _qv = dmu.embed_text(query)
            _cv = dmu.embed_text(response_text) if response_text else None
            if _qv and _cv:
                chunk_vec = _cv
                sim_Q = round(dmu.calculate_similarity_vec(_qv, _cv, "Cosine"), 3)
        except Exception as _e:
            logger.warning(f"FunctionSearch chunk embedding failed ({func_name}): {_e}")

        if chunk_vec is not None:
            rag_selected.append({
                "rag_name":  rag.get("RAG_NAME", ""),
                "rag":       rag.get("RAG_NAME", ""),
                "bucket":    "FunctionSearch",
                "DB":        "FunctionSearch",
                "id":        func_name,
                "ID":        func_name,
                "title":     func_name,
                "function_name": func_name,
                "query":     query,
                "args":      input_str,
                "response":  response_text,
                "category":  func_name,
                "value_text": response_text,
                "vector_data_value_text": chunk_vec,
                "similarity_prompt": sim_Q,
                "query_mode": "NORMAL",
                "QUERY_MODE": "NORMAL",
                "query_seq":  "0",
                "QUERY_SEQ":  "0",
                "log_format": log_template,
            })
        else:
            try:
                _legacy = rag.get("LOG_TEMPLATE",
                    "'rag':'{rag_name}', 'function':'{function_name}'").format(
                        rag_name=rag.get("RAG_NAME", ""),
                        function_name=func_name)
            except (KeyError, IndexError):
                _legacy = f"rag={rag.get('RAG_NAME', '')}, function={func_name}"
            rag_selected.append(_legacy)

        exec_info.setdefault("_FUNCTION_SEARCH_LOG", []).append({
            "rag_name": rag.get("RAG_NAME", ""),
            "function_name": func_name,
            "query": query,
            "args": input_str,
            "response": response_text,
            "timestamp": ts_begin,
        })

    if rag_context:
        header = rag.get("HEADER_TEMPLATE", "")
        rag_context = header + rag_context
        text_limits = rag.get("TEXT_LIMITS", 6000)
        if len(rag_context) > text_limits:
            rag_context = rag_context[:text_limits]
    return rag_context, rag_selected


# Build context from a PageIndex RAG
def select_rag_page_index(query, rag, exec_info):
    rag_context = ""
    rag_selected = []

    # Resolve the page folder from DATA_NAME
    for rd in rag.get("DATA", []):
        if rd.get("DATA_TYPE") != "PAGE_INDEX":
            continue
        data_name = rd["DATA_NAME"]
        pages_dir = str(Path(rag_folder_pages_path) / data_name)
        index_path = str(Path(pages_dir) / "_index.json")

        if not os.path.exists(index_path):
            logger.warning(f"Page index not found: {index_path}")
            continue

        # Load the index
        index_data = dmu.read_json_file(index_path)
        pages = index_data.get("PAGES", [])
        if not pages:
            continue

        # Map for looking up meta by page id
        pages_map = {p["id"]: p for p in pages}

        # Let the LLM choose page IDs
        max_pages = rag.get("MAX_PAGES", 5)
        support_agent = rd.get("SUPPORT_AGENT", "agent_59PageIndexSearch.json")
        selected_ids = dmt.page_index_search(
            exec_info, support_agent, query, pages, max_pages)

        # Load the selected page bodies and assemble dict entries that
        # `get_knowledge_reference` can score (so each page gets a contribution
        # metric in Analytics Result - Knowledge Utility).
        log_template = rag.get("LOG_TEMPLATE",
            "'rag':'{rag_name}', 'page_id':'{page_id}', 'title':'{title}', "
            "'category':'{category}', 'summary':'{summary}'")
        # Append analytics-required fields (similarity_Q/A/DB/ID/QUERY_*)
        # to the user-defined template if missing.
        log_template = _augment_log_template(log_template)
        # Embed the user query once per call so per-page similarity is cheap.
        try:
            query_vec = dmu.embed_text(query)
        except Exception as _e:
            logger.warning(f"PageIndex query embedding failed: {_e}")
            query_vec = None
        for page_id in selected_ids:
            md_path = str(Path(pages_dir) / f"{page_id}.md")
            if not os.path.exists(md_path):
                continue
            page_content = dmu.read_text_file(md_path)
            _crumb = _build_page_breadcrumb(page_id, pages_map)
            if _crumb:
                rag_context += f"[Path] {_crumb}\n\n{page_content}\n\n"
            else:
                rag_context += page_content + "\n\n"
            page_meta = pages_map.get(page_id, {})

            # Compute similarity_prompt against the chunk text (page body).
            chunk_vec, sim_Q = None, 0.0
            try:
                if query_vec:
                    chunk_vec = dmu.embed_text(page_content)
                    if chunk_vec:
                        sim_Q = round(
                            dmu.calculate_similarity_vec(query_vec, chunk_vec, "Cosine"), 3)
            except Exception as _e:
                logger.warning(f"PageIndex chunk embedding failed (id={page_id}): {_e}")
                chunk_vec, sim_Q = None, 0.0

            if chunk_vec is not None:
                rag_selected.append({
                    "rag_name":  rag.get("RAG_NAME", ""),
                    "rag":       rag.get("RAG_NAME", ""),
                    "bucket":    "PageIndex",
                    "DB":        "PageIndex",
                    "id":        str(page_id),
                    "ID":        str(page_id),
                    "page_id":   str(page_id),
                    "title":     page_meta.get("title", ""),
                    "category":  page_meta.get("category", ""),
                    "summary":   page_meta.get("summary", ""),
                    "value_text": page_content,
                    "vector_data_value_text": chunk_vec,
                    "similarity_prompt": sim_Q,
                    "query_mode": "NORMAL",
                    "QUERY_MODE": "NORMAL",
                    "query_seq":  "0",
                    "QUERY_SEQ":  "0",
                    "log_format": log_template,
                })
            else:
                # Embedding unavailable → fall back to a legacy string log entry
                # (chunk is still in context, just absent from analytics).
                _legacy_tpl = rag.get("LOG_TEMPLATE",
                    "'rag':'{rag_name}', 'page_id':'{page_id}', 'title':'{title}', "
                    "'category':'{category}', 'summary':'{summary}'")
                try:
                    _legacy = _legacy_tpl.format(
                        rag_name=rag.get("RAG_NAME", ""),
                        page_id=str(page_id),
                        title=page_meta.get("title", ""),
                        category=page_meta.get("category", ""),
                        summary=page_meta.get("summary", ""))
                except (KeyError, IndexError):
                    _legacy = (f"rag={rag.get('RAG_NAME', '')}, page_id={page_id}, "
                               f"title={page_meta.get('title', '')}")
                rag_selected.append(_legacy)

    # Prepend the header template
    if rag_context:
        header = rag.get("HEADER_TEMPLATE", "")
        rag_context = header + rag_context

        # Truncate at the text cap
        text_limits = rag.get("TEXT_LIMITS", 6000)
        if len(rag_context) > text_limits:
            rag_context = rag_context[:text_limits]

    return rag_context, rag_selected

# Generate RAG data
def generate_rag():
    # Load the RAG master
    rag_mst_dict = dmu.read_json_file(rag_mst_file, mst_folder_path)

    # Generate each RAG data
    cnt_add = 0
    cnt_extent = 0

    # Fetch chunk data
    rag_data = []
    for rag_id, rag_setting in rag_mst_dict.items():
        if rag_setting["active"] == "Y":
            if rag_setting["input"] == "notion":
                if rag_setting.get("data_type") == "pageindex":
                    rag_data = get_chunk_notion_pageindex(rag_setting["bucket"], rag_setting["data_name"], rag_setting["item_dict"], rag_setting.get("chk_dict"), rag_setting.get("date_dict"), rag_setting.get("category_dict"))
                else:
                    rag_data = get_chunk_notion(rag_setting["bucket"], rag_setting["data_name"], rag_setting["item_dict"], rag_setting["chk_dict"], rag_setting["date_dict"], rag_setting["category_dict"])
            elif rag_setting["input"] == "csv":
                if isinstance(rag_setting["file_name"], list):
                    for rag_data_file_name in rag_setting["file_name"]:
                        rag_data += get_chunk_csv(rag_setting["bucket"], rag_setting["file_path"], rag_data_file_name, rag_setting["field_items"], rag_setting["title"], rag_setting["key_text"], rag_setting["value_text"], rag_setting["category_items"])
                else:
                    rag_data = get_chunk_csv(rag_setting["bucket"], rag_setting["file_path"], rag_setting["file_name"], rag_setting["field_items"], rag_setting["title"], rag_setting["key_text"], rag_setting["value_text"], rag_setting["category_items"])
            elif rag_setting["input"] == "excel":
                if rag_setting.get("data_type") == "pageindex":
                    rag_data = get_chunk_excel_pageindex(
                        rag_setting["bucket"],
                        rag_setting["source_dir"],
                        rag_setting["source_file"],
                        rag_setting.get("sheet", "pages"),
                        rag_setting["item_dict"],
                    )
                else:
                    logger.warning("Excel input currently supports only pageindex")
            else:
                logger.warning("No valid mode configured.")

            if rag_data:
                # Keep the page IDs (copy before save_rag_chunk_db removes id)
                page_ids_map = {chunk["id"]: chunk["id"] for chunk in rag_data if "id" in chunk}

                # Save into ChromaDB
                if rag_setting["data_type"] == "chromadb":
                    cnt_add, cnt_extent = save_rag_chunk_db(rag_id, rag_data)
                    cnt_total = cnt_add + cnt_extent
                    logger.info(f"{rag_id} DB write done. added: {cnt_add}, total: {cnt_total}")

                    # Reflect the RAG-registration complete flag back to Notion (targets pages without fin_flg, so update all)
                    fin_flg = rag_setting.get("fin_flg", {})
                    if fin_flg and rag_setting["input"] == "notion":
                        fin_cnt = 0
                        for page_id in page_ids_map.values():
                            for prop_name, prop_value in fin_flg.items():
                                try:
                                    if isinstance(prop_value, bool):
                                        dmn.update_notion_chk(page_id, prop_name, prop_value)
                                    else:
                                        logger.warning(f"fin_flg: unsupported type {prop_name}={prop_value}")
                                    fin_cnt += 1
                                except Exception as e:
                                    logger.warning(f"fin_flg update failed (page={page_id}, {prop_name}): {e}")
                        logger.info(f"{rag_id}: fin_flg reflected to {fin_cnt} Notion pages")

                # Save into PageIndex (Notion-derived)
                elif rag_setting["data_type"] == "pageindex":
                    processed = save_rag_pageindex(rag_setting["bucket"], rag_data)
                    logger.info(f"{rag_id} PageIndex write done. count: {len(processed)}")

                    # Reflect the RAG-registration complete flag back to Notion (Notion input only)
                    fin_flg = rag_setting.get("fin_flg", {})
                    if fin_flg and rag_setting["input"] == "notion":
                        fin_cnt = 0
                        for rec in processed:
                            notion_pid = rec.get("notion_page_id")
                            if not notion_pid:
                                continue
                            for prop_name, prop_value in fin_flg.items():
                                try:
                                    if isinstance(prop_value, bool):
                                        dmn.update_notion_chk(notion_pid, prop_name, prop_value)
                                    else:
                                        logger.warning(f"fin_flg: unsupported type {prop_name}={prop_value}")
                                    fin_cnt += 1
                                except Exception as e:
                                    logger.warning(f"fin_flg update failed (page={notion_pid}, {prop_name}): {e}")
                        logger.info(f"{rag_id}: fin_flg reflected to {fin_cnt} Notion pages")

                # Legacy PageIndex (CSV-style chunks with page_id)
                elif rag_setting["data_type"] == "page_index":
                    cnt_add, cnt_total = save_rag_pages(rag_id, rag_data)
                    logger.info(f"{rag_id} page write done. pages: {cnt_total}")

                    # Reflect the RAG-registration complete flag back to Notion
                    fin_flg = rag_setting.get("fin_flg", {})
                    if fin_flg and rag_setting["input"] == "notion":
                        fin_cnt = 0
                        for page_id in page_ids_map.values():
                            for prop_name, prop_value in fin_flg.items():
                                try:
                                    if isinstance(prop_value, bool):
                                        dmn.update_notion_chk(page_id, prop_name, prop_value)
                                    else:
                                        logger.warning(f"fin_flg: unsupported type {prop_name}={prop_value}")
                                    fin_cnt += 1
                                except Exception as e:
                                    logger.warning(f"fin_flg update failed (page={page_id}, {prop_name}): {e}")
                        logger.info(f"{rag_id}: fin_flg reflected to {fin_cnt} Notion pages")

# Delete a RAG database (Collection)
def del_rag_db(ragdb_selected=[]):
    db_client = get_chroma_client()

    # Connect to SQLite3
    conn = sqlite3.connect(rag_folder_db_path+'chroma.sqlite3')

    if ragdb_selected:
        # Fetch the dataframe of rows to delete
        placeholders = ','.join(['?'] * len(ragdb_selected))  # '?,?,?'
        query = f"SELECT id AS collection_id, name FROM collections WHERE name IN ({placeholders})"
        collections_df = pd.read_sql_query(query, conn, params=ragdb_selected)
        segments_df = pd.read_sql_query("SELECT id AS segment_id, collection AS collection_id , type, scope FROM segments WHERE scope = 'VECTOR'", conn)
        merged_df = pd.merge(segments_df, collections_df, on="collection_id", how="inner")

        for collection_name in ragdb_selected:
            db_client.delete_collection(name=collection_name)
            target_segments = merged_df[merged_df["name"] == collection_name]["segment_id"].tolist()
            for segment_id in target_segments:
                seg_path = Path(rag_folder_db_path) / segment_id
                if seg_path.exists():
                    shutil.rmtree(seg_path)
                    logger.info(f"Deleted: {seg_path}")
                else:
                    logger.warning(f"Not found: {seg_path}")
            logger.info(f"Deleted {collection_name}.")
    else:
        # Fetch the dataframe of rows to delete
        query = "SELECT id AS collection_id, name FROM collections"
        collections_df = pd.read_sql_query(query, conn, params=ragdb_selected)
        segments_df = pd.read_sql_query("SELECT id AS segment_id, collection AS collection_id , type, scope FROM segments WHERE scope = 'VECTOR'", conn)
        merged_df = pd.merge(segments_df, collections_df, on="collection_id", how="inner")

        collections = db_client.list_collections()
        for collection in collections:
            collection_name = collection.name
            db_client.delete_collection(name=collection_name)
            target_segments = merged_df[merged_df["name"] == collection_name]["segment_id"].tolist()
            for segment_id in target_segments:
                seg_path = Path(rag_folder_db_path) / segment_id
                if seg_path.exists():
                    shutil.rmtree(seg_path)
                    logger.info(f"Deleted: {seg_path}")
                else:
                    logger.warning(f"Not found: {seg_path}")
            logger.info(f"Deleted {collection_name}.")

    # Release SQLite3 physical space
    conn.execute("VACUUM;")
    conn.close()

    # Reset the singleton (re-connect required after deletion)
    global _chroma_client
    _chroma_client = None
